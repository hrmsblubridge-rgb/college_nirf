from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
import uuid
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from datetime import datetime, timezone
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from college_data import COLLEGES
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Models ───────────────────────────────────────────────────────────────────
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

# ── Helpers ──────────────────────────────────────────────────────────────────
def normalize(s: str) -> str:
    return s.lower().strip() if s else ""

def match_college(name: str, colleges: list) -> Optional[dict]:
    """Match college name (full or short) against DB. Returns college doc or None."""
    if not name or not name.strip():
        return None
    n = normalize(name)
    # 1. Exact match on full name
    for c in colleges:
        if normalize(c["college_name"]) == n:
            return c
    # 2. Exact match on any short name
    for c in colleges:
        for s in c.get("short_names", []):
            if normalize(s) == n:
                return c
    # 3. Fuzzy match (threshold 0.82)
    best, best_score = None, 0
    for c in colleges:
        candidates = [c["college_name"]] + c.get("short_names", [])
        for cand in candidates:
            score = SequenceMatcher(None, n, normalize(cand)).ratio()
            if score > best_score:
                best_score = score
                best = c
    if best_score >= 0.82:
        return best
    return None

def rank_display(c: Optional[dict]) -> str:
    """Return rank string. 'NL' (Not Listed) if no match found."""
    if c is None:
        return "NL"
    r = c.get("rank")
    if r is None:
        return "NL"
    return str(r)

def excel_thin_border():
    t = Side(style="thin", color="D1D5DB")
    return Border(left=t, right=t, top=t, bottom=t)

# ── Basic Routes ─────────────────────────────────────────────────────────────
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    return status_checks

# ── College DB Routes ─────────────────────────────────────────────────────────
@api_router.post("/colleges/seed")
async def seed_colleges():
    """Seed all 300 colleges into MongoDB."""
    await db.colleges.drop()
    await db.colleges.create_index("college_name")
    result = await db.colleges.insert_many(COLLEGES)
    return {"seeded": len(result.inserted_ids), "message": "College database ready"}

@api_router.get("/colleges")
async def get_colleges(skip: int = 0, limit: int = 300):
    colleges = await db.colleges.find({}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.colleges.count_documents({})
    return {"colleges": colleges, "total": total}

@api_router.get("/colleges/stats")
async def college_stats():
    total = await db.colleges.count_documents({})
    ranked = await db.colleges.count_documents({"rank": {"$type": "int"}})
    rank_band = await db.colleges.count_documents({"rank": {"$type": "string"}})
    return {"total": total, "ranked": ranked, "unranked": rank_band}

# ── Excel Processing Route ────────────────────────────────────────────────────
@api_router.post("/process-excel")
async def process_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Please upload an Excel file (.xlsx or .xls)")

    # Load colleges from DB
    colleges = await db.colleges.find({}, {"_id": 0}).to_list(500)
    if not colleges:
        colleges = COLLEGES  # fallback to in-memory

    # Read uploaded Excel
    content = await file.read()
    wb_in = openpyxl.load_workbook(io.BytesIO(content))
    ws_in = wb_in.active

    # Find column indices (1-based)
    headers = [ws_in.cell(row=1, column=c).value for c in range(1, ws_in.max_column + 1)]
    headers_lower = [str(h).lower().strip() if h else "" for h in headers]

    def find_col(keywords):
        for kw in keywords:
            for i, h in enumerate(headers_lower):
                if kw in h:
                    return i + 1  # 1-based
        return None

    ug_col  = find_col(["ug university", "ug uni"])
    pg_col  = find_col(["pg university", "pg uni"])

    # Detect if a "Rank" column already exists right after UG university col
    # If col ug_col+1 header contains "rank" → use it; otherwise we'll INSERT a new column
    ug_has_rank_col = False
    if ug_col:
        next_hdr = str(headers[ug_col] if ug_col < len(headers) else "").lower()  # 0-indexed
        if "rank" in next_hdr:
            ug_has_rank_col = True

    pg_has_rank_col = False
    if pg_col:
        next_hdr_pg = str(headers[pg_col] if pg_col < len(headers) else "").lower()
        if "rank" in next_hdr_pg:
            pg_has_rank_col = True

    # ── Build output workbook ──────────────────────────────────────────────
    wb_out = openpyxl.Workbook()

    # ─ Sheet 1: Build headers list — insert "Rank" if missing ─────────────
    ws1 = wb_out.active
    ws1.title = "Applicant Data with Ranks"

    hdr_fill  = PatternFill("solid", fgColor="1A73E8")
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Calibri", size=10)
    rank_font = Font(name="Calibri", bold=True, color="1557B0", size=10)
    null_font = Font(name="Calibri", italic=True, color="9CA3AF", size=10)
    band_font = Font(name="Calibri", bold=True, color="E67E22", size=10)
    nl_font   = Font(name="Calibri", italic=True, color="DC2626", size=10)
    alt_fill  = PatternFill("solid", fgColor="F0F7FF")

    # Build new headers list (insert Rank columns where missing)
    new_headers = list(headers)
    ug_rank_out_col = None   # 1-based column in OUTPUT
    pg_rank_out_col = None

    insert_ug = 0  # how many columns were inserted before PG col
    if ug_col and not ug_has_rank_col:
        # Insert "Rank" column right after ug_col
        new_headers.insert(ug_col, "Rank")  # ug_col is 0-indexed insert position (after index ug_col-1)
        insert_ug = 1
        ug_rank_out_col = ug_col + 1  # 1-based: original ug_col is still ug_col, rank is ug_col+1
    elif ug_col and ug_has_rank_col:
        ug_rank_out_col = ug_col + 1  # already exists

    if pg_col and not pg_has_rank_col:
        pg_col_adj = pg_col + insert_ug  # adjusted for any UG insertion
        new_headers.insert(pg_col_adj, "Rank")  # insert rank after pg university col
        pg_rank_out_col = pg_col_adj + 1
    elif pg_col and pg_has_rank_col:
        pg_rank_out_col = pg_col + insert_ug + 1

    # Write headers to ws1
    for c, h in enumerate(new_headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = excel_thin_border()
    ws1.row_dimensions[1].height = 28

    # Track original input names + matched rank for Tab 2
    # List of (original_input_name, rank_string) — one per applicant row
    ug_tab2_rows = []   # [(original_ug_name, rank_str), ...]

    # Copy rows + fill ranks, mapping input columns → output columns (with possible rank col insertions)
    for row_idx in range(2, ws_in.max_row + 1):
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else None

        out_col = 1
        for in_col in range(1, ws_in.max_column + 1):
            # If we inserted a UG rank column, skip that output slot when copying regular data
            if ug_col and not ug_has_rank_col and out_col == ug_col + 1:
                out_col += 1  # leave slot for inserted UG Rank
            if pg_col and not pg_has_rank_col and out_col == pg_col + insert_ug + 1:
                out_col += 1  # leave slot for inserted PG Rank

            val = ws_in.cell(row=row_idx, column=in_col).value
            cell = ws1.cell(row=row_idx, column=out_col, value=val)
            cell.font = data_font
            if row_fill:
                cell.fill = row_fill
            cell.border = excel_thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            out_col += 1

        ws1.row_dimensions[row_idx].height = 16

        # Fill UG rank in the correct output column
        if ug_col and ug_rank_out_col:
            ug_name = ws_in.cell(row=row_idx, column=ug_col).value
            if ug_name and str(ug_name).strip():
                original_name = str(ug_name).strip()
                matched = match_college(original_name, colleges)
                rank_val = rank_display(matched)
                ug_tab2_rows.append((original_name, rank_val))
                if rank_val:
                    r_cell = ws1.cell(row=row_idx, column=ug_rank_out_col)
                    r_cell.value = rank_val
                    if rank_val == "NL":
                        r_cell.font = nl_font
                    elif rank_val.startswith(("101-", "151-", "201-")):
                        r_cell.font = band_font
                    else:
                        r_cell.font = rank_font
                    r_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fill PG rank in the correct output column
        if pg_col and pg_rank_out_col:
            pg_name = ws_in.cell(row=row_idx, column=pg_col).value
            if pg_name and str(pg_name).strip():
                matched = match_college(str(pg_name), colleges)
                rank_val = rank_display(matched)
                if rank_val:
                    r_cell = ws1.cell(row=row_idx, column=pg_rank_out_col)
                    r_cell.value = rank_val
                    if rank_val == "NL":
                        r_cell.font = nl_font
                    elif rank_val.startswith(("101-", "151-", "201-")):
                        r_cell.font = band_font
                    else:
                        r_cell.font = rank_font
                    r_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-width for first few key columns
    for col in ws1.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    ws1.freeze_panes = "A2"

    # ─ Sheet 2: Ranking Reference ──────────────────────────────────────────
    ws2 = wb_out.create_sheet("Ranking Reference")

    # Column widths
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 14

    # Header row — matches screenshot exactly: white bg, bold text, centered
    hdr_cells = [("UG University/institute Name", 1), ("Rank", 2)]
    for label, c in hdr_cells:
        cell = ws2.cell(row=1, column=c, value=label)
        cell.font = Font(name="Calibri", bold=True, size=11, color="111827")
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.border = excel_thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[1].height = 36

    # Data rows — exact input names in order, alternating light green fill
    green_fill  = PatternFill("solid", fgColor="E8F5E9")   # light green
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    for i, (orig_name, rank_str) in enumerate(ug_tab2_rows):
        row_num = i + 2
        row_fill = green_fill  # all rows get the light green (matches screenshot)

        # Name cell
        cell_name = ws2.cell(row=row_num, column=1, value=orig_name)
        cell_name.font = Font(name="Calibri", size=11, color="111827")
        cell_name.fill = row_fill
        cell_name.border = excel_thin_border()
        cell_name.alignment = Alignment(horizontal="center", vertical="center")

        # Rank cell
        rank_val_display = rank_str if rank_str else ""
        cell_rank = ws2.cell(row=row_num, column=2, value=rank_val_display)
        cell_rank.font = Font(name="Calibri", size=11, color="111827")
        cell_rank.fill = row_fill
        cell_rank.border = excel_thin_border()
        cell_rank.alignment = Alignment(horizontal="center", vertical="center")

        ws2.row_dimensions[row_num].height = 28

    ws2.freeze_panes = "A2"

    # ── Stream back as download ──────────────────────────────────────────────
    out_buffer = io.BytesIO()
    wb_out.save(out_buffer)
    out_buffer.seek(0)

    return StreamingResponse(
        out_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="BluBridge_Processed_{file.filename}"'}
    )

# ── Static Excel Downloads ────────────────────────────────────────────────────
@api_router.get("/download-college-list")
async def download_college_list():
    file_path = Path("/app/college_rankings.xlsx")
    if not file_path.exists():
        raise HTTPException(404, "File not found")
    return FileResponse(str(file_path),
        filename="BluBridge_College_Rankings_NIRF_2025.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api_router.get("/download-college-list-shortnames")
async def download_college_list_shortnames():
    file_path = Path("/app/college_rankings_with_shortnames.xlsx")
    if not file_path.exists():
        raise HTTPException(404, "File not found")
    return FileResponse(str(file_path),
        filename="BluBridge_College_Rankings_ShortNames_NIRF_2025.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── All-India Colleges Routes ─────────────────────────────────────────────────
@api_router.post("/all-india-colleges/seed")
async def seed_all_india_colleges():
    """Seed all ~37K Indian colleges into MongoDB with NIRF rank matching."""
    json_path = Path("/tmp/all_india_colleges.json")
    if not json_path.exists():
        raise HTTPException(404, "College data file not found. Generate it first.")
    with open(json_path) as f:
        records = json.load(f)
    await db.all_india_colleges.drop()
    await db.all_india_colleges.create_index("college_name")
    await db.all_india_colleges.create_index("rank")
    # Insert in batches
    batch_size = 5000
    total = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        result = await db.all_india_colleges.insert_many(batch)
        total += len(result.inserted_ids)
    return {"seeded": total, "message": f"{total} Indian colleges loaded into database"}

@api_router.get("/all-india-colleges/stats")
async def all_india_stats():
    total = await db.all_india_colleges.count_documents({})
    ranked = await db.all_india_colleges.count_documents({"rank": {"$ne": "NL"}})
    return {"total": total, "ranked": ranked, "nl": total - ranked}

@api_router.get("/download-all-india-colleges")
async def download_all_india_colleges():
    file_path = Path("/app/frontend/public/indian_colleges_sorted.xlsx")
    if not file_path.exists():
        raise HTTPException(404, "File not found")
    return FileResponse(str(file_path),
        filename="All_India_Colleges_with_Ranks.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Excel Matcher: Match two Excel sheets by phone number ─────────────────────
import re as _re
import pandas as pd

def normalize_phone(val):
    """Extract last 10 digits from phone number for matching."""
    if pd.isna(val) or val is None:
        return ''
    s = _re.sub(r'[^0-9]', '', str(val).strip())
    return s[-10:] if len(s) >= 10 else s

@api_router.post("/excel-matcher/preview")
async def excel_matcher_preview(
    left_file: UploadFile = File(...),
    right_file: UploadFile = File(...)
):
    """Preview columns from both uploaded Excel files."""
    try:
        left_bytes = await left_file.read()
        right_bytes = await right_file.read()

        left_df = pd.read_excel(io.BytesIO(left_bytes))
        right_df = pd.read_excel(io.BytesIO(right_bytes))

        # Save to temp for later processing
        left_path = Path("/tmp/matcher_left.xlsx")
        right_path = Path("/tmp/matcher_right.xlsx")
        left_path.write_bytes(left_bytes)
        right_path.write_bytes(right_bytes)

        return {
            "left_columns": list(left_df.columns),
            "right_columns": list(right_df.columns),
            "left_rows": len(left_df),
            "right_rows": len(right_df),
            "left_sample": left_df.head(3).fillna('').to_dict(orient='records'),
            "right_sample": right_df.head(3).fillna('').to_dict(orient='records'),
        }
    except Exception as e:
        raise HTTPException(400, f"Error reading Excel files: {str(e)}")


@api_router.post("/excel-matcher/process")
async def excel_matcher_process(
    left_phone_col: str,
    right_phone_col: str,
    right_status_col: str,
):
    """Match left sheet with right sheet by phone, add Shortlist/Reject column, return file."""
    import xlsxwriter as _xw

    left_path = Path("/tmp/matcher_left.xlsx")
    right_path = Path("/tmp/matcher_right.xlsx")

    if not left_path.exists() or not right_path.exists():
        raise HTTPException(400, "Please upload both files first via /excel-matcher/preview")

    left_df = pd.read_excel(left_path)
    right_df = pd.read_excel(right_path)

    if left_phone_col not in left_df.columns:
        raise HTTPException(400, f"Column '{left_phone_col}' not found in left file")
    if right_phone_col not in right_df.columns:
        raise HTTPException(400, f"Column '{right_phone_col}' not found in right file")
    if right_status_col not in right_df.columns:
        raise HTTPException(400, f"Column '{right_status_col}' not found in right file")

    # Build phone→status lookup from right sheet
    phone_status = {}
    for _, row in right_df.iterrows():
        phone = normalize_phone(row.get(right_phone_col))
        status = str(row.get(right_status_col, '')).strip()
        if phone and status:
            phone_status[phone] = status

    # Match and add column to left sheet
    statuses = []
    matched = 0
    for _, row in left_df.iterrows():
        phone = normalize_phone(row.get(left_phone_col))
        status = phone_status.get(phone, '')
        if status:
            matched += 1
        statuses.append(status)

    # Insert "Shortlist / Reject" column right after the phone column
    phone_col_idx = list(left_df.columns).index(left_phone_col)
    left_df.insert(phone_col_idx + 1, 'Shortlist / Reject', statuses)

    # Write output with xlsxwriter
    output_path = Path("/tmp/matcher_result.xlsx")
    wb = _xw.Workbook(str(output_path), {'strings_to_numbers': False})
    ws = wb.add_worksheet('Matched Results')

    # Formats
    hdr_fmt = wb.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#1A73E8',
        'font_size': 11, 'font_name': 'Calibri', 'align': 'center', 'valign': 'vcenter',
        'text_wrap': True, 'border': 1})
    data_fmt = wb.add_format({'font_size': 10, 'font_name': 'Calibri', 'border': 1,
        'border_color': '#D1D5DB', 'valign': 'vcenter'})
    data_alt = wb.add_format({'font_size': 10, 'font_name': 'Calibri', 'border': 1,
        'border_color': '#D1D5DB', 'valign': 'vcenter', 'bg_color': '#F0F7FF'})
    shortlist_fmt = wb.add_format({'font_size': 10, 'font_name': 'Calibri', 'border': 1,
        'bold': True, 'font_color': '#16A34A', 'bg_color': '#DCFCE7', 'align': 'center', 'valign': 'vcenter'})
    reject_fmt = wb.add_format({'font_size': 10, 'font_name': 'Calibri', 'border': 1,
        'bold': True, 'font_color': '#DC2626', 'bg_color': '#FEE2E2', 'align': 'center', 'valign': 'vcenter'})
    empty_fmt = wb.add_format({'font_size': 10, 'font_name': 'Calibri', 'border': 1,
        'font_color': '#9CA3AF', 'align': 'center', 'valign': 'vcenter', 'italic': True})

    # Headers
    status_col_idx = phone_col_idx + 1
    for ci, col in enumerate(left_df.columns):
        ws.write(0, ci, str(col), hdr_fmt)
        ws.set_column(ci, ci, max(15, len(str(col)) + 4))
    ws.set_row(0, 28)

    # Data
    for ri, (_, row) in enumerate(left_df.iterrows()):
        r = ri + 1
        alt = r % 2 == 0
        for ci, col in enumerate(left_df.columns):
            val = row[col]
            cell_val = '' if pd.isna(val) else str(val)

            if ci == status_col_idx:
                # Status column - special formatting
                low = cell_val.lower().strip()
                if 'shortlist' in low:
                    ws.write(r, ci, cell_val, shortlist_fmt)
                elif 'reject' in low:
                    ws.write(r, ci, cell_val, reject_fmt)
                else:
                    ws.write(r, ci, cell_val if cell_val else 'Not Found', empty_fmt)
            else:
                ws.write(r, ci, cell_val, data_alt if alt else data_fmt)

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, len(left_df), len(left_df.columns) - 1)
    wb.close()

    return FileResponse(
        str(output_path),
        filename="Matched_Results.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"X-Matched": str(matched), "X-Total": str(len(left_df)),
                 "X-Unmatched": str(len(left_df) - matched)}
    )


# ── App Setup ─────────────────────────────────────────────────────────────────
app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

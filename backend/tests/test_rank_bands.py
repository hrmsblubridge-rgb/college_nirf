"""
Test suite for BluBridge College Ranking Processor - Rank Band Feature
Tests:
1. GET /api/colleges/stats - stats verification
2. GET /api/colleges with rank bands (RB:101-150, RB:151-200, RB:201-300)
3. POST /api/process-excel - Excel processing with rank bands and NL (Not Listed)
4. GET /api/download-college-list - static download verification
5. GET /api/download-college-list-shortnames - static download with shortnames
"""

import pytest
import requests
import os
import io
import openpyxl
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestCollegeStats:
    """Test /api/colleges/stats endpoint"""

    def test_stats_endpoint_returns_expected_values(self):
        """Stats should return total:300, ranked:100 (numeric), unranked:200 (band strings)"""
        response = requests.get(f"{BASE_URL}/api/colleges/stats")
        assert response.status_code == 200, f"Stats endpoint failed: {response.text}"
        
        data = response.json()
        assert data["total"] == 300, f"Expected total=300, got {data['total']}"
        assert data["ranked"] == 100, f"Expected ranked=100 (numeric), got {data['ranked']}"
        assert data["unranked"] == 200, f"Expected unranked=200 (rank bands), got {data['unranked']}"


class TestCollegesWithRankBands:
    """Test /api/colleges endpoint with rank band colleges"""

    def test_colleges_skip_100_returns_rb_101_150(self):
        """Skip 100 should return colleges with rank 'RB:101-150'"""
        response = requests.get(f"{BASE_URL}/api/colleges?skip=100&limit=5")
        assert response.status_code == 200, f"Colleges endpoint failed: {response.text}"
        
        data = response.json()
        assert len(data["colleges"]) == 5, f"Expected 5 colleges, got {len(data['colleges'])}"
        
        # All should have RB:101-150 rank
        for college in data["colleges"]:
            assert college["rank"] == "RB:101-150", f"Expected rank 'RB:101-150', got '{college['rank']}' for {college['college_name']}"

    def test_colleges_skip_160_returns_rb_151_200(self):
        """Skip 160 should return colleges with rank 'RB:151-200'"""
        response = requests.get(f"{BASE_URL}/api/colleges?skip=160&limit=5")
        assert response.status_code == 200, f"Colleges endpoint failed: {response.text}"
        
        data = response.json()
        assert len(data["colleges"]) == 5, f"Expected 5 colleges, got {len(data['colleges'])}"
        
        # All should have RB:151-200 rank
        for college in data["colleges"]:
            assert college["rank"] == "RB:151-200", f"Expected rank 'RB:151-200', got '{college['rank']}' for {college['college_name']}"

    def test_colleges_skip_215_returns_rb_201_300(self):
        """Skip 215 should return colleges with rank 'RB:201-300'"""
        response = requests.get(f"{BASE_URL}/api/colleges?skip=215&limit=5")
        assert response.status_code == 200, f"Colleges endpoint failed: {response.text}"
        
        data = response.json()
        assert len(data["colleges"]) == 5, f"Expected 5 colleges, got {len(data['colleges'])}"
        
        # All should have RB:201-300 rank
        for college in data["colleges"]:
            assert college["rank"] == "RB:201-300", f"Expected rank 'RB:201-300', got '{college['rank']}' for {college['college_name']}"

    def test_colleges_skip_0_returns_numeric_ranks(self):
        """Skip 0 should return top ranked colleges with numeric ranks 1-5"""
        response = requests.get(f"{BASE_URL}/api/colleges?skip=0&limit=5")
        assert response.status_code == 200, f"Colleges endpoint failed: {response.text}"
        
        data = response.json()
        assert len(data["colleges"]) == 5
        
        # All should have numeric ranks
        for college in data["colleges"]:
            rank = college["rank"]
            assert isinstance(rank, int), f"Expected numeric rank, got '{rank}' (type {type(rank)}) for {college['college_name']}"


class TestProcessExcelWithRankBands:
    """Test /api/process-excel with mixed rank types and unknown colleges"""

    def test_process_excel_mixed_colleges(self):
        """
        Upload Excel with:
        - IIT Madras (rank 1)
        - Anurag University (RB:101-150)
        - Tezpur University (RB:151-200)
        - Reva University (RB:201-300)
        - XYZ Unknown College (NL - Not Listed)
        """
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Applicants"
        
        # Headers
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="UG University")
        
        # Test data with various rank types
        test_colleges = [
            ("Student A", "Indian Institute of Technology Madras"),  # Rank 1
            ("Student B", "IIT Madras"),  # Short name - Rank 1
            ("Student C", "Anurag University"),  # RB:101-150
            ("Student D", "Tezpur University"),  # RB:151-200
            ("Student E", "Reva University"),  # RB:201-300
            ("Student F", "XYZ Unknown College"),  # NL (Not Listed)
            ("Student G", "REVA"),  # Short name - RB:201-300
        ]
        
        for i, (name, college) in enumerate(test_colleges, start=2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=college)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Upload to API
        files = {'file': ('test_applicants.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/process-excel", files=files)
        
        assert response.status_code == 200, f"Process Excel failed: {response.text}"
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('content-type', '')
        
        # Parse response Excel
        result_wb = openpyxl.load_workbook(io.BytesIO(response.content))
        result_ws = result_wb.active
        
        # Find the Rank column
        rank_col = None
        for col in range(1, result_ws.max_column + 1):
            if result_ws.cell(row=1, column=col).value == "Rank":
                rank_col = col
                break
        
        assert rank_col is not None, "Rank column not found in output"
        
        # Expected ranks for each row
        expected_ranks = {
            2: "1",           # IIT Madras - numeric rank
            3: "1",           # IIT Madras short name
            4: "RB:101-150",  # Anurag University
            5: "RB:151-200",  # Tezpur University
            6: "RB:201-300",  # Reva University
            7: "NL",          # XYZ Unknown College
            8: "RB:201-300",  # REVA short name
        }
        
        for row, expected in expected_ranks.items():
            actual = str(result_ws.cell(row=row, column=rank_col).value or "")
            assert actual == expected, f"Row {row}: Expected rank '{expected}', got '{actual}'"
        
        print("All ranks matched correctly in processed Excel!")


class TestStaticDownloads:
    """Test static Excel download endpoints with rank bands"""

    def test_download_college_list_has_rank_bands(self):
        """Download college list should contain rank bands instead of null"""
        response = requests.get(f"{BASE_URL}/api/download-college-list")
        assert response.status_code == 200, f"Download failed: {response.text}"
        
        # Parse Excel
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find rank column
        rank_col = None
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col).value or "").lower()
            if "rank" in header:
                rank_col = col
                break
        
        assert rank_col is not None, "Rank column not found in downloaded Excel"
        
        # Check that we have numeric ranks and rank bands (no nulls)
        numeric_count = 0
        band_101_150_count = 0
        band_151_200_count = 0
        band_201_300_count = 0
        null_count = 0
        
        for row in range(2, ws.max_row + 1):
            rank = ws.cell(row=row, column=rank_col).value
            if rank is None or str(rank).strip() == "":
                null_count += 1
            elif isinstance(rank, int) or str(rank).isdigit():
                numeric_count += 1
            elif str(rank) == "RB:101-150":
                band_101_150_count += 1
            elif str(rank) == "RB:151-200":
                band_151_200_count += 1
            elif str(rank) == "RB:201-300":
                band_201_300_count += 1
        
        assert null_count == 0, f"Found {null_count} null/empty ranks - should be rank bands!"
        assert numeric_count == 100, f"Expected 100 numeric ranks, got {numeric_count}"
        assert band_101_150_count >= 40, f"Expected ~50 RB:101-150 bands, got {band_101_150_count}"
        assert band_151_200_count >= 40, f"Expected ~50 RB:151-200 bands, got {band_151_200_count}"
        assert band_201_300_count >= 80, f"Expected ~100 RB:201-300 bands, got {band_201_300_count}"
        
        print(f"Download verified: {numeric_count} numeric, {band_101_150_count} RB:101-150, {band_151_200_count} RB:151-200, {band_201_300_count} RB:201-300")

    def test_download_college_list_shortnames_has_rank_bands(self):
        """Download college list with shortnames should contain rank bands"""
        response = requests.get(f"{BASE_URL}/api/download-college-list-shortnames")
        assert response.status_code == 200, f"Download failed: {response.text}"
        
        # Parse Excel
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find rank column
        rank_col = None
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col).value or "").lower()
            if "rank" in header:
                rank_col = col
                break
        
        assert rank_col is not None, "Rank column not found in downloaded Excel"
        
        # Check for null/empty ranks
        null_count = 0
        band_count = 0
        
        for row in range(2, ws.max_row + 1):
            rank = ws.cell(row=row, column=rank_col).value
            if rank is None or str(rank).strip() == "":
                null_count += 1
            elif str(rank).startswith("RB:"):
                band_count += 1
        
        assert null_count == 0, f"Found {null_count} null/empty ranks - should be rank bands!"
        assert band_count >= 180, f"Expected ~200 rank band entries, got {band_count}"
        
        print(f"Shortnames download verified: {band_count} rank bands, {null_count} nulls")


class TestBasicEndpoints:
    """Basic API health checks"""

    def test_root_endpoint(self):
        """Root endpoint should respond"""
        response = requests.get(f"{BASE_URL}/api/")
        assert response.status_code == 200

    def test_colleges_total_count(self):
        """Total college count should be 300"""
        response = requests.get(f"{BASE_URL}/api/colleges?limit=1")
        assert response.status_code == 200
        data = response.json()
        assert data["total"] == 300


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])

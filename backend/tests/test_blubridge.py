"""BluBridge backend API tests"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestHealth:
    def test_root(self):
        r = requests.get(f"{BASE_URL}/api/")
        assert r.status_code == 200


class TestCollegeStats:
    def test_stats_returns_300(self):
        r = requests.get(f"{BASE_URL}/api/colleges/stats")
        assert r.status_code == 200
        data = r.json()
        assert data["total"] == 300, f"Expected 300, got {data['total']}"
        assert data["ranked"] == 100, f"Expected 100 ranked, got {data['ranked']}"
        assert "unranked" in data

    def test_colleges_list(self):
        r = requests.get(f"{BASE_URL}/api/colleges")
        assert r.status_code == 200
        data = r.json()
        assert data["total"] == 300


class TestSeedDB:
    def test_seed_endpoint(self):
        r = requests.post(f"{BASE_URL}/api/colleges/seed")
        assert r.status_code == 200
        data = r.json()
        assert data["seeded"] == 300


class TestProcessExcel:
    """Test excel processing with rank matching"""

    def test_process_excel_returns_xlsx(self):
        with open("/tmp/sample-data.xlsx", "rb") as f:
            r = requests.post(
                f"{BASE_URL}/api/process-excel",
                files={"file": ("sample-data.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_process_excel_has_two_sheets(self):
        import openpyxl
        with open("/tmp/sample-data.xlsx", "rb") as f:
            r = requests.post(
                f"{BASE_URL}/api/process-excel",
                files={"file": ("sample-data.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) == 2, f"Expected 2 sheets, got {wb.sheetnames}"
        print(f"Sheet names: {wb.sheetnames}")

    def test_vit_rank_16_full_name(self):
        """VIT full name should get rank 16"""
        import openpyxl
        # Create minimal excel with VIT
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["UG University/institute Name", "Rank UG"])
        ws.append(["Vellore Institute of Technology", ""])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        r = requests.post(
            f"{BASE_URL}/api/process-excel",
            files={"file": ("test.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert r.status_code == 200
        wb_out = openpyxl.load_workbook(io.BytesIO(r.content))
        ws1 = wb_out.active
        rank_val = ws1.cell(row=2, column=2).value
        assert str(rank_val) == "16", f"VIT full name rank expected 16, got {rank_val}"

    def test_vit_rank_16_short_name(self):
        """VIT short name should also get rank 16"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["UG University/institute Name", "Rank UG"])
        ws.append(["VIT", ""])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        r = requests.post(
            f"{BASE_URL}/api/process-excel",
            files={"file": ("test.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert r.status_code == 200
        wb_out = openpyxl.load_workbook(io.BytesIO(r.content))
        rank_val = wb_out.active.cell(row=2, column=2).value
        assert str(rank_val) == "16", f"VIT short name rank expected 16, got {rank_val}"

    def test_jadavpur_rank_18(self):
        """Jadavpur University (JU) should get rank 18"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["UG University/institute Name", "Rank UG"])
        ws.append(["Jadavpur University", ""])
        ws.append(["JU", ""])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        r = requests.post(
            f"{BASE_URL}/api/process-excel",
            files={"file": ("test.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert r.status_code == 200
        wb_out = openpyxl.load_workbook(io.BytesIO(r.content))
        ws1 = wb_out.active
        rank_ju_full = ws1.cell(row=2, column=2).value
        rank_ju_short = ws1.cell(row=3, column=2).value
        assert str(rank_ju_full) == "18", f"Jadavpur full name rank expected 18, got {rank_ju_full}"
        assert str(rank_ju_short) == "18", f"JU short name rank expected 18, got {rank_ju_short}"

    def test_rank_band_college_returns_null(self):
        """Rank-band colleges (rank=None) should return 'null'"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["UG University/institute Name", "Rank UG"])
        # From college_data, rank=None colleges
        ws.append(["Bansilal Ramnath Agarwal Charitable Trust's Vishwakarma Institute of Technology", ""])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        r = requests.post(
            f"{BASE_URL}/api/process-excel",
            files={"file": ("test.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert r.status_code == 200
        wb_out = openpyxl.load_workbook(io.BytesIO(r.content))
        rank_val = wb_out.active.cell(row=2, column=2).value
        assert str(rank_val) == "null", f"Rank-band college expected 'null', got {rank_val}"

    def test_invalid_file_returns_400(self):
        r = requests.post(
            f"{BASE_URL}/api/process-excel",
            files={"file": ("test.txt", b"not excel", "text/plain")}
        )
        assert r.status_code == 400


class TestDownloads:
    def test_download_college_list(self):
        r = requests.get(f"{BASE_URL}/api/download-college-list")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_download_college_list_shortnames(self):
        r = requests.get(f"{BASE_URL}/api/download-college-list-shortnames")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

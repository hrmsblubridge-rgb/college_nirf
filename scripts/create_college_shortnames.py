import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════
# SHORT NAME RULES:
# - Only include GLOBALLY/NATIONALLY recognized abbreviations
# - If college is known by TWO names, format: "NAME1 / NAME2"
# - Empty string "" if no globally recognized short name
# ═══════════════════════════════════════════════════════════════════════

# ── TOP 100 RANKED  ── format: (rank, full_name, short_name, city, state)
ranked = [
    (1,  "Indian Institute of Technology Madras",                           "IIT Madras",                     "Chennai",               "Tamil Nadu"),
    (2,  "Indian Institute of Technology Delhi",                            "IIT Delhi",                      "New Delhi",             "Delhi"),
    (3,  "Indian Institute of Technology Bombay",                           "IIT Bombay",                     "Mumbai",                "Maharashtra"),
    (4,  "Indian Institute of Technology Kanpur",                           "IIT Kanpur",                     "Kanpur",                "Uttar Pradesh"),
    (5,  "Indian Institute of Technology Kharagpur",                        "IIT Kharagpur / IIT KGP",        "Kharagpur",             "West Bengal"),
    (6,  "Indian Institute of Technology Roorkee",                          "IIT Roorkee / IITR",             "Roorkee",               "Uttarakhand"),
    (7,  "Indian Institute of Technology Hyderabad",                        "IIT Hyderabad / IITH",           "Hyderabad",             "Telangana"),
    (8,  "Indian Institute of Technology Guwahati",                         "IIT Guwahati / IITG",            "Guwahati",              "Assam"),
    (9,  "National Institute of Technology Tiruchirappalli",                 "NIT Trichy / NITT",              "Tiruchirappalli",       "Tamil Nadu"),
    (10, "Indian Institute of Technology (Banaras Hindu University) Varanasi","IIT BHU / IIT (BHU)",          "Varanasi",              "Uttar Pradesh"),
    (11, "Birla Institute of Technology & Science - Pilani",                 "BITS Pilani / BITS",             "Pilani",                "Rajasthan"),
    (12, "Indian Institute of Technology Indore",                           "IIT Indore",                     "Indore",                "Madhya Pradesh"),
    (13, "National Institute of Technology Rourkela",                       "NIT Rourkela / NITR",            "Rourkela",              "Odisha"),
    (14, "S.R.M. Institute of Science and Technology",                      "SRM / SRMIST",                   "Chennai",               "Tamil Nadu"),
    (15, "Indian Institute of Technology (Indian School of Mines)",         "IIT ISM / IIT (ISM)",            "Dhanbad",               "Jharkhand"),
    (16, "Vellore Institute of Technology",                                 "VIT",                            "Vellore",               "Tamil Nadu"),
    (17, "National Institute of Technology Karnataka, Surathkal",           "NITK / NIT Surathkal",           "Surathkal",             "Karnataka"),
    (18, "Jadavpur University",                                             "JU",                             "Kolkata",               "West Bengal"),
    (19, "Indian Institute of Technology Patna",                            "IIT Patna",                      "Patna",                 "Bihar"),
    (20, "Anna University",                                                 "Anna University",                "Chennai",               "Tamil Nadu"),
    (21, "National Institute of Technology Calicut",                        "NIT Calicut / NITC",             "Kozhikode",             "Kerala"),
    (22, "Siksha 'O' Anusandhan",                                           "SOA University",                 "Bhubaneswar",           "Odisha"),
    (23, "Amrita Vishwavidyapeetham",                                       "Amrita",                         "Coimbatore",            "Tamil Nadu"),
    (24, "Jamia Millia Islamia",                                            "JMI",                            "New Delhi",             "Delhi"),
    (25, "Indian Institute of Technology Gandhinagar",                      "IIT Gandhinagar / IITGN",        "Gandhinagar",           "Gujarat"),
    (26, "Indian Institute of Technology Mandi",                            "IIT Mandi",                      "Mandi",                 "Himachal Pradesh"),
    (27, "Indian Institute of Technology Jodhpur",                          "IIT Jodhpur",                    "Jodhpur",               "Rajasthan"),
    (28, "National Institute of Technology Warangal",                       "NITW / NIT Warangal",            "Warangal",              "Telangana"),
    (29, "Thapar Institute of Engineering and Technology",                  "TIET / Thapar",                  "Patiala",               "Punjab"),
    (30, "Delhi Technological University",                                  "DTU",                            "New Delhi",             "Delhi"),
    (31, "Chandigarh University",                                           "CU",                             "Mohali",                "Punjab"),
    (32, "Indian Institute of Technology Ropar",                            "IIT Ropar",                      "Rupnagar",              "Punjab"),
    (33, "Kalasalingam Academy of Research and Education",                  "",                               "Krishnan Koil",         "Tamil Nadu"),
    (34, "Aligarh Muslim University",                                       "AMU",                            "Aligarh",               "Uttar Pradesh"),
    (35, "Koneru Lakshmaiah Education Foundation University (K L College of Engineering)", "KL University / KLU", "Vaddeswaram",       "Andhra Pradesh"),
    (36, "Kalinga Institute of Industrial Technology",                      "KIIT",                           "Bhubaneswar",           "Odisha"),
    (37, "Amity University",                                                "Amity",                          "Gautam Budh Nagar",     "Uttar Pradesh"),
    (38, "International Institute of Information Technology Hyderabad",     "IIIT Hyderabad / IIIT-H",        "Hyderabad",             "Telangana"),
    (39, "Indian Institute of Technology Bhubaneswar",                      "IIT Bhubaneswar",                "Bhubaneswar",           "Odisha"),
    (40, "Shanmugha Arts Science Technology & Research Academy",            "SASTRA",                         "Thanjavur",             "Tamil Nadu"),
    (41, "Institute of Chemical Technology",                                "ICT Mumbai",                     "Mumbai",                "Maharashtra"),
    (42, "Manipal Academy of Higher Education",                             "MAHE / Manipal",                 "Manipal",               "Karnataka"),
    (43, "UPES",                                                            "UPES",                           "Dehradun",              "Uttarakhand"),
    (44, "Visvesvaraya National Institute of Technology, Nagpur",           "VNIT / VNIT Nagpur",             "Nagpur",                "Maharashtra"),
    (45, "Saveetha Institute of Medical and Technical Sciences",            "SIMATS",                         "Chennai",               "Tamil Nadu"),
    (46, "Symbiosis International",                                         "SIU",                            "Pune",                  "Maharashtra"),
    (47, "Sri Sivasubramaniya Nadar College of Engineering",                "SSN",                            "Kalavakkam",            "Tamil Nadu"),
    (48, "Lovely Professional University",                                  "LPU",                            "Phagwara",              "Punjab"),
    (49, "National Institute of Technology Durgapur",                       "NIT Durgapur",                   "Durgapur",              "West Bengal"),
    (50, "National Institute of Technology Silchar",                        "NIT Silchar / NITS",             "Silchar",               "Assam"),
    (51, "Birla Institute of Technology",                                   "BIT Mesra / BIT",                "Ranchi",                "Jharkhand"),
    (52, "Graphic Era University",                                          "",                               "Dehradun",              "Uttarakhand"),
    (53, "National Institute of Technology Patna",                          "NIT Patna",                      "Patna",                 "Bihar"),
    (54, "Indian Institute of Engineering Science and Technology, Shibpur", "IIEST Shibpur / IIEST",          "Howrah",                "West Bengal"),
    (55, "Dr. B.R Ambedkar National Institute of Technology, Jalandhar",   "NIT Jalandhar / NITJ",           "Jalandhar",             "Punjab"),
    (56, "Indian Institute of Technology Jammu",                            "IIT Jammu",                      "Jammu",                 "Jammu and Kashmir"),
    (57, "Indian Institute of Technology, Tirupati",                        "IIT Tirupati",                   "Tirupati",              "Andhra Pradesh"),
    (58, "Manipal University Jaipur",                                       "MUJ",                            "Jaipur",                "Rajasthan"),
    (59, "Manipal Institute of Technology",                                 "MIT Manipal",                    "Manipal",               "Karnataka"),
    (60, "Madan Mohan Malaviya University of Technology",                   "MMMUT",                          "Gorakhpur",             "Uttar Pradesh"),
    (61, "Indian Institute of Space Science and Technology",                "IIST",                           "Thiruvananthapuram",    "Kerala"),
    (62, "Motilal Nehru National Institute of Technology",                  "MNNIT",                          "Prayagraj",             "Uttar Pradesh"),
    (63, "Indraprastha Institute of Information Technology",                "IIIT Delhi / IIIT-D",            "New Delhi",             "Delhi"),
    (64, "Indian Institute of Technology Palakkad",                         "IIT Palakkad",                   "Palakkad",              "Kerala"),
    (65, "National Institute of Technology Delhi",                          "NIT Delhi",                      "Delhi",                 "Delhi"),
    (66, "Sardar Vallabhbhai National Institute of Technology",             "SVNIT / NIT Surat",              "Surat",                 "Gujarat"),
    (67, "PSG College of Technology",                                       "PSG Tech",                       "Coimbatore",            "Tamil Nadu"),
    (68, "National Institute of Technology, Andhra Pradesh",               "NIT Andhra Pradesh",             "Tadepalligudem",        "Andhra Pradesh"),
    (69, "Sathyabama Institute of Science and Technology",                  "Sathyabama",                     "Chennai",               "Tamil Nadu"),
    (70, "International Institute of Information Technology Bangalore",     "IIIT Bangalore / IIIT-B",        "Bengaluru",             "Karnataka"),
    (71, "Netaji Subhas University of Technology (NSUT)",                   "NSUT",                           "Delhi",                 "Delhi"),
    (72, "Banasthali Vidyapith",                                            "",                               "Banasthali",            "Rajasthan"),
    (73, "Indian Institute of Technology Bhilai",                           "IIT Bhilai",                     "Durg",                  "Chhattisgarh"),
    (74, "National Institute of Technology Srinagar",                       "NIT Srinagar",                   "Srinagar",              "Jammu and Kashmir"),
    (75, "University of Hyderabad",                                         "UoH",                            "Hyderabad",             "Telangana"),
    (76, "M. S. Ramaiah Institute of Technology",                           "MSRIT",                          "Bengaluru",             "Karnataka"),
    (77, "Christ University",                                               "Christ",                         "Bengaluru",             "Karnataka"),
    (78, "Indian Institute of Technology Dharwad",                          "IIT Dharwad",                    "Dharwad",               "Karnataka"),
    (79, "Rajiv Gandhi Institute of Petroleum Technology",                  "RGIPT",                          "Amethi",                "Uttar Pradesh"),
    (80, "Sant Longowal Institute of Engineering & Technology",             "SLIET",                          "Longowal",              "Punjab"),
    (81, "Vignan's Foundation for Science, Technology and Research",        "VFSTR / Vignan's",               "Guntur",                "Andhra Pradesh"),
    (82, "Maulana Azad National Institute of Technology",                   "MANIT / MANIT Bhopal",           "Bhopal",                "Madhya Pradesh"),
    (83, "National Institute of Technology, Jamshedpur",                   "NIT Jamshedpur",                 "Jamshedpur",            "Jharkhand"),
    (84, "National Institute of Technology Meghalaya",                      "NIT Meghalaya",                  "Shillong",              "Meghalaya"),
    (85, "Jain University, Bangalore",                                      "",                               "Bengaluru",             "Karnataka"),
    (86, "National Institute of Technology Kurukshetra",                    "NIT Kurukshetra / NITKKR",       "Kurukshetra",           "Haryana"),
    (87, "National Institute of Technology, Raipur",                        "NIT Raipur",                     "Raipur",                "Chhattisgarh"),
    (88, "Vel Tech Rangarajan Dr. Sagunthala R&D Institute of Science and Technology", "Vel Tech",           "Chennai",               "Tamil Nadu"),
    (89, "AU College of Engineering (A)",                                   "AUCE",                           "Visakhapatnam",         "Andhra Pradesh"),
    (90, "Chitkara University",                                             "",                               "Rajpura",               "Punjab"),
    (91, "COEP Technological University",                                   "COEP",                           "Pune",                  "Maharashtra"),
    (92, "SR University",                                                   "",                               "Warangal",              "Telangana"),
    (93, "Defence Institute of Advanced Technology",                        "DIAT",                           "Pune",                  "Maharashtra"),
    (94, "Panjab University",                                               "PU",                             "Chandigarh",            "Chandigarh"),
    (95, "Jawaharlal Nehru Technological University",                       "JNTUH",                          "Hyderabad",             "Telangana"),
    (96, "C.V. Raman Global University, Odisha",                            "",                               "Bhubaneswar",           "Odisha"),
    (97, "Atal Bihari Vajpayee Indian Institute of Information Technology and Management", "ABV-IIITM / IIITM Gwalior", "Gwalior",   "Madhya Pradesh"),
    (98, "National Institute of Technology Hamirpur",                       "NIT Hamirpur",                   "Hamirpur",              "Himachal Pradesh"),
    (99, "Pandit Deendayal Energy University",                              "PDEU",                           "Gandhinagar",           "Gujarat"),
    (100,"National Institute of Technology Puducherry",                     "NIT Puducherry",                 "Karaikal",              "Puducherry"),
]

# ── BAND 101–150  ── format: (full_name, short_name, city, state)
band_101_150 = [
    ("Amity University",                                                    "Amity",          "North Twenty Four Parganas",  "West Bengal"),
    ("Amity University Haryana, Gurgaon",                                   "",               "Gurugram",                    "Haryana"),
    ("Anurag University",                                                   "",               "Hyderabad",                   "Telangana"),
    ("Bansilal Ramnath Agarwal Charitable Trust's Vishwakarma Institute of Technology", "VIT Pune", "Pune",                 "Maharashtra"),
    ("Chandigarh Engineering College-CGC, Landran, Mohali",                 "CGC",            "Mohali",                      "Punjab"),
    ("Chennai Institute of Technology",                                     "",               "Chennai",                     "Tamil Nadu"),
    ("Coimbatore Institute of Technology",                                  "CIT",            "Coimbatore",                  "Tamil Nadu"),
    ("College of Engineering Trivandrum",                                   "CET",            "Thiruvananthapuram",          "Kerala"),
    ("Dr. Vishwanath Karad MIT World Peace University",                     "MIT-WPU",        "Pune",                        "Maharashtra"),
    ("Easwari Engineering College",                                         "",               "Chennai",                     "Tamil Nadu"),
    ("Galgotias University",                                                "",               "Gautam Budh Nagar",           "Uttar Pradesh"),
    ("Gandhi Institute of Technology And Management (GITAM)",               "GITAM",          "Visakhapatnam",               "Andhra Pradesh"),
    ("Guru Gobind Singh Indraprastha University",                           "GGSIPU / IPU",   "New Delhi",                   "Delhi"),
    ("Hindustan Institute of Technology and Science (HITS)",                "HITS",           "Chennai",                     "Tamil Nadu"),
    ("Indian Institute of Information Technology Allahabad",                "IIIT Allahabad / IIITA", "Prayagraj",           "Uttar Pradesh"),
    ("Indian Institute of Technology Goa",                                  "IIT Goa",        "Ponda",                       "Goa"),
    ("Jawaharlal Nehru Technological University",                           "JNTUK",          "Kakinada",                    "Andhra Pradesh"),
    ("Jaypee Institute of Information Technology",                          "JIIT",           "Noida",                       "Uttar Pradesh"),
    ("K L University",                                                      "KLU",            "Guntur",                      "Andhra Pradesh"),
    ("Karunya Institute of Technology and Sciences",                        "Karunya",        "Coimbatore",                  "Tamil Nadu"),
    ("Kongu Engineering College",                                           "",               "Perundurai",                  "Tamil Nadu"),
    ("KPR Institute of Engineering and Technology",                         "",               "Coimbatore",                  "Tamil Nadu"),
    ("Maharishi Markandeshwar",                                             "",               "Ambala",                      "Haryana"),
    ("Mahindra University",                                                 "",               "Hyderabad",                   "Telangana"),
    ("Manav Rachna International Institute of Research & Studies",          "",               "Faridabad",                   "Haryana"),
    ("Maulana Azad National Urdu University",                               "MANUU",          "Hyderabad",                   "Telangana"),
    ("National Institute for Solar Energy",                                 "NISE",           "Gurugram",                    "Haryana"),
    ("National Institute of Food Technology, Entrepreneurship & Management","NIFTEM",         "Sonepat",                     "Haryana"),
    ("Nirma University",                                                    "Nirma",          "Ahmedabad",                   "Gujarat"),
    ("Nitte Meenakshi Institute of Technology",                             "NMIT",           "Bengaluru",                   "Karnataka"),
    ("Noida Institute of Engineering & Technology",                         "",               "Greater Noida",               "Uttar Pradesh"),
    ("Pandit Dwarka Prasad Mishra IIITDM Jabalpur",                         "PDPM IIITDM / IIITDM Jabalpur", "Jabalpur",   "Madhya Pradesh"),
    ("PES University",                                                      "PESU",           "Bengaluru",                   "Karnataka"),
    ("PSG Institute of Technology and Applied Research",                    "",               "Coimbatore",                  "Tamil Nadu"),
    ("Punjab Engineering College (Deemed to be University), Chandigarh",    "PEC",            "Chandigarh",                  "Chandigarh"),
    ("R.V. College of Engineering",                                         "RVCE",           "Bengaluru",                   "Karnataka"),
    ("Rajalakshmi Engineering College",                                     "",               "Chennai",                     "Tamil Nadu"),
    ("Sharda University",                                                   "",               "Greater Noida",               "Uttar Pradesh"),
    ("Shoolini University of Biotechnology and Management Sciences",        "",               "Solan",                       "Himachal Pradesh"),
    ("Siddaganga Institute of Technology",                                  "SIT",            "Tumkur",                      "Karnataka"),
    ("SVKM's Narsee Monjee Institute of Management Studies",                "NMIMS",          "Mumbai",                      "Maharashtra"),
    ("The Northcap University",                                             "NCU",            "Gurugram",                    "Haryana"),
    ("Thiagarajar College of Engineering",                                  "TCE",            "Madurai",                     "Tamil Nadu"),
    ("University College of Engineering",                                   "",               "Hyderabad",                   "Telangana"),
    ("University of Allahabad",                                             "",               "Prayagraj",                   "Uttar Pradesh"),
    ("Veermata Jijabai Technological Institute (VJTI, Mumbai)",             "VJTI",           "Mumbai",                      "Maharashtra"),
    ("National Institute of Technology Agartala",                           "NIT Agartala / NITA", "Agartala",              "Tripura"),
    ("National Institute of Technology Arunachal Pradesh",                  "NIT Arunachal",  "Itanagar",                    "Arunachal Pradesh"),
    ("National Institute of Technology Goa",                                "NIT Goa",        "Cuncolim",                    "Goa"),
    ("National Institute of Technology Mizoram",                            "NIT Mizoram",    "Aizawl",                      "Mizoram"),
]

# ── BAND 151–200  ──
band_151_200 = [
    ("Amity University Rajasthan, Jaipur",                                  "",               "Jaipur",                      "Rajasthan"),
    ("B. S. Abdur Rahman Crescent Institute of Science and Technology",     "",               "Chennai",                     "Tamil Nadu"),
    ("B.M.S. College of Engineering",                                       "BMSCE",          "Bengaluru",                   "Karnataka"),
    ("Bharati Vidyapeeth Deemed University College of Engineering",         "",               "Pune",                        "Maharashtra"),
    ("C M R Institute of Technology",                                       "CMRIT",          "Bengaluru",                   "Karnataka"),
    ("CGC College of Engineering, Landran",                                 "CGC",            "Sahibzada Ajit Singh Nagar",  "Punjab"),
    ("Chaitanya Bharathi Institute of Technology",                          "CBIT",           "Hyderabad",                   "Telangana"),
    ("CVR College of Engineering",                                          "",               "Ibrahimpatam",                "Telangana"),
    ("Dayalbagh Educational Institute",                                     "",               "Agra",                        "Uttar Pradesh"),
    ("Dr. D. Y. Patil Institute of Technology",                             "",               "Pune",                        "Maharashtra"),
    ("Dr. M. G. R. Educational and Research Institute",                     "",               "Chennai",                     "Tamil Nadu"),
    ("G. L. A. University",                                                 "",               "Mathura",                     "Uttar Pradesh"),
    ("G.L. Bajaj Institute of Technology and Management",                   "",               "Greater Noida",               "Uttar Pradesh"),
    ("Goka Raju Ranga Raju Institute of Engineering & Technology",          "",               "Hyderabad",                   "Telangana"),
    ("Guru Ghasidas Vishwavidyalaya",                                       "",               "Bilaspur",                    "Chhattisgarh"),
    ("Indian Institute of Information Technology, Design & Manufacturing, Kancheepuram", "IIITDM Kancheepuram", "Chennai",  "Tamil Nadu"),
    ("Institute of Aeronautical Engineering",                               "",               "Hyderabad",                   "Telangana"),
    ("Institute of Engineering & Management",                               "",               "Kolkata",                     "West Bengal"),
    ("Integral University",                                                 "",               "Lucknow",                     "Uttar Pradesh"),
    ("Islamic University of Science & Technology, Pulwama",                 "",               "Pulwama",                     "Jammu and Kashmir"),
    ("J. C. Bose University of Science and Technology, YMCA",              "",               "Faridabad",                   "Haryana"),
    ("Jawaharlal Nehru Technological University, Ananthapuramu",            "JNTUA",          "Ananthapuramu",               "Andhra Pradesh"),
    ("JSPM's Rajarshi Shahu College of Engineering",                        "",               "Pune",                        "Maharashtra"),
    ("KIET Group of Institutions",                                          "KIET",           "Ghaziabad",                   "Uttar Pradesh"),
    ("KLE Technological University",                                        "KLE Tech",       "Dharwad",                     "Karnataka"),
    ("Kumaraguru College of Technology",                                    "KCT",            "Coimbatore",                  "Tamil Nadu"),
    ("Maulana Abul Kalam Azad University of Technology",                    "MAKAUT / WBUT",  "Nadia",                       "West Bengal"),
    ("Mepco Schlenk Engineering College",                                   "",               "Sivakasi",                    "Tamil Nadu"),
    ("National Institute of Technology Manipur",                            "NIT Manipur",    "Imphal",                      "Manipur"),
    ("National Institute of Technology Sikkim",                             "NIT Sikkim",     "South Sikkim",                "Sikkim"),
    ("National Institute of Technology Uttarakhand",                        "NIT Uttarakhand","Srinagar (Garhwal)",          "Uttarakhand"),
    ("National Institute of Technology Nagaland",                           "NIT Nagaland",   "Dimapur",                     "Nagaland"),
    ("New Horizon College of Engineering",                                  "",               "Bengaluru",                   "Karnataka"),
    ("NMAM Institute of Technology",                                        "NMAMIT",         "Nitte, Udupi",                "Karnataka"),
    ("North Eastern Regional Institute of Science & Technology",            "NERIST",         "Itanagar",                    "Arunachal Pradesh"),
    ("R.M.K. Engineering College",                                          "",               "Thiruvallur",                 "Tamil Nadu"),
    ("Ramdeobaba University, Nagpur",                                       "RCOEM",          "Nagpur",                      "Maharashtra"),
    ("Shri Mata Vaishno Devi University",                                   "SMVDU",          "Katra",                       "Jammu and Kashmir"),
    ("Sona College of Technology",                                          "",               "Salem",                       "Tamil Nadu"),
    ("Sri Ramakrishna Engineering College",                                 "SREC",           "Coimbatore",                  "Tamil Nadu"),
    ("Sri Sai Ram Institute of Technology",                                 "",               "Chennai",                     "Tamil Nadu"),
    ("Sri Sairam Engineering College",                                      "",               "Kancheepuram",                "Tamil Nadu"),
    ("Tezpur University",                                                   "",               "Tezpur",                      "Assam"),
    ("The National Institute of Engineering",                               "NIE Mysore",     "Mysuru",                      "Karnataka"),
    ("Vallurupalli Nageswara Rao Vignana Jyothi Institute of Engineering and Technology", "VNR VJIET", "Hyderabad",         "Telangana"),
    ("Vardhaman College of Engineering",                                    "",               "Rangareddy",                  "Telangana"),
    ("Veer Surendra Sai University of Technology",                          "VSSUT",          "Sambalpur",                   "Odisha"),
    ("Velagapudi Ramakrishna Siddhartha Engineering College",               "VRSEC",          "Vijayawada",                  "Andhra Pradesh"),
    ("Vels Institute of Science Technology and Advanced Studies (VISTAS)",  "VISTAS",         "Chennai",                     "Tamil Nadu"),
    ("Vignan Institute of Technology and Science",                          "",               "Yadadri-Bhuvanagiri",         "Telangana"),
    ("Visvesvaraya Technological University",                               "VTU",            "Belgaum",                     "Karnataka"),
    ("Yeshwantrao Chavan College of Engineering",                           "YCCE",           "Nagpur",                      "Maharashtra"),
]

# ── BAND 201–300  ──
band_201_300 = [
    ("Aditya Institute of Technology and Management",                       "",               "Tekkali",                     "Andhra Pradesh"),
    ("Aditya University",                                                   "",               "Surampalem",                  "Andhra Pradesh"),
    ("Ajeenkya D Y Patil University",                                       "",               "Pune",                        "Maharashtra"),
    ("Amity University Patna",                                              "",               "Patna",                       "Bihar"),
    ("Amity University, Gwalior",                                           "",               "Gwalior",                     "Madhya Pradesh"),
    ("Amity University, Jharkhand",                                         "",               "Ranchi",                      "Jharkhand"),
    ("Annamalai University",                                                "",               "Annamalainagar",              "Tamil Nadu"),
    ("Army Institute of Technology",                                        "AIT",            "Pune",                        "Maharashtra"),
    ("B I T SINDRI",                                                        "BIT Sindri",     "Dhanbad",                     "Jharkhand"),
    ("BML Munjal University",                                               "BMU",            "Gurgaon",                     "Haryana"),
    ("BMS Institute of Technology & Management",                            "BMSIT&M",        "Bengaluru",                   "Karnataka"),
    ("Centurion University of Technology and Management",                   "CUTM",           "Paralakhemundi",              "Odisha"),
    ("Chandigarh Engineering College Jhanjeri",                             "",               "Sahibzada Ajit Singh Nagar",  "Punjab"),
    ("CMR College of Engineering & Technology",                             "",               "Hyderabad",                   "Telangana"),
    ("CMR Technical Campus",                                                "",               "Hyderabad",                   "Telangana"),
    ("College of Engineering & Technology, Bhubaneswar",                   "",               "Bhubaneswar",                 "Odisha"),
    ("Dayananda Sagar College of Engineering",                              "DSCE",           "Bengaluru",                   "Karnataka"),
    ("Dhirubhai Ambani Institute of Information and Communication Technology", "DA-IICT / DAIICT", "Gandhinagar",           "Gujarat"),
    ("Dr. Shyama Prasad Mukherjee International Institute of Information Technology, Naya Raipur", "IIIT Naya Raipur", "Raipur", "Chhattisgarh"),
    ("E.G.S. Pillay Engineering College",                                   "",               "Nagapattinam",                "Tamil Nadu"),
    ("G. H. Raisoni College of Engineering, Nagpur",                        "GHRCE",          "Nagpur",                      "Maharashtra"),
    ("G. Narayanamma Institute of Technology & Science for Women",          "GNITS",          "Hyderabad",                   "Telangana"),
    ("Galgotias College of Engineering & Technology",                       "",               "Greater Noida",               "Uttar Pradesh"),
    ("Gandhi Institute for Technological Advancement (GITA), Bhubaneswar", "GITA",           "Bhubaneswar",                 "Odisha"),
    ("GIET University, Gunupur",                                            "GIET",           "Gunupur",                     "Odisha"),
    ("GMR Institute of Technology",                                         "",               "Rajahmundry",                 "Andhra Pradesh"),
    ("Godavari Institute of Engineering & Technology",                      "",               "Rajahmundry",                 "Andhra Pradesh"),
    ("Guru Jambheshwar University of Science and Technology, Hisar",        "GJUST",          "Hisar",                       "Haryana"),
    ("Harcourt Butler Technical University",                                "HBTU",           "Kanpur Nagar",                "Uttar Pradesh"),
    ("Hindusthan College of Engineering and Technology",                    "",               "Coimbatore",                  "Tamil Nadu"),
    ("IES College of Technology, Bhopal",                                   "",               "Bhopal",                      "Madhya Pradesh"),
    ("Indian Institute of Information Technology Guwahati",                 "IIIT Guwahati",  "Guwahati",                    "Assam"),
    ("Indian Institute of Petroleum & Energy",                              "IIPE",           "Visakhapatnam",               "Andhra Pradesh"),
    ("Indira Gandhi Delhi Technical University for Women",                  "IGDTUW",         "Delhi",                       "Delhi"),
    ("Jamia Hamdard",                                                       "",               "New Delhi",                   "Delhi"),
    ("Jaypee University of Information Technology",                         "JUIT",           "Solan",                       "Himachal Pradesh"),
    ("JIS College of Engineering",                                          "",               "Kalyani",                     "West Bengal"),
    ("JSS Academy Of Technical Education, Noida",                           "",               "Gautam Budh Nagar",           "Uttar Pradesh"),
    ("JSS Science and Technology University",                               "JSSTU",          "Mysuru",                      "Karnataka"),
    ("K. Ramakrishnan College of Engineering",                              "",               "Samayapuram",                 "Tamil Nadu"),
    ("K. Ramakrishnan College of Technology",                               "",               "Tiruchirappalli",             "Tamil Nadu"),
    ("Kakatiya Institute of Technology & Science",                          "KITS Warangal",  "Warangal",                    "Telangana"),
    ("Kalaignar Karunanidhi Institute of Technology",                       "",               "Coimbatore",                  "Tamil Nadu"),
    ("Karpagam College of Engineering",                                     "",               "Coimbatore",                  "Tamil Nadu"),
    ("Lakshmi Narain College of Technology",                                "LNCT",           "Bhopal",                      "Madhya Pradesh"),
    ("M. Kumarasamy College of Engineering",                                "",               "Karur",                       "Tamil Nadu"),
    ("Madanapalle Institute of Technology & Science",                       "MITS",           "Madanapalle",                 "Andhra Pradesh"),
    ("Maharaja Sayajirao University of Baroda",                             "MSU Baroda / MSUB", "Vadodara",                 "Gujarat"),
    ("Maharshi Dayanand University, Rohtak",                                "MDU",            "Rohtak",                      "Haryana"),
    ("Maharshi Karve Stree Shikshan Samstha's Cummins College of Engineering for Women", "COEW / Cummins", "Pune",           "Maharashtra"),
    ("Mahatma Jyotiba Phule Rohilkhand University, Bareilly",               "MJP Rohilkhand", "Bareilly",                   "Uttar Pradesh"),
    ("Malla Reddy Engineering College",                                     "MREC",           "Hyderabad",                   "Telangana"),
    ("Malla Reddy Engineering College for Women (Autonomous)",              "MRECW",          "Hyderabad",                   "Telangana"),
    ("Manav Rachna University",                                             "MRU",            "Faridabad",                   "Haryana"),
    ("Marwadi University",                                                  "",               "Rajkot",                      "Gujarat"),
    ("MIT Art, Design and Technology University, Pune",                     "MIT ADT",        "Pune",                        "Maharashtra"),
    ("MLR Institute of Technology",                                         "",               "Hyderabad",                   "Telangana"),
    ("Narula Institute of Technology",                                      "NIT Kolkata",    "Kolkata",                     "West Bengal"),
    ("National Engineering College",                                        "NEC",            "Kovilpatti",                  "Tamil Nadu"),
    ("National Institute of Advanced Manufacturing Technology, Ranchi",     "NIAMT",          "Ranchi",                      "Jharkhand"),
    ("P E S College of Engineering, MANDYA",                                "PESCE",          "Mandya",                      "Karnataka"),
    ("Padmashree Dr. D.Y. Patil Vidyapeeth, Mumbai",                        "",               "Mumbai",                      "Maharashtra"),
    ("Panimalar Engineering College",                                       "",               "Thiruvallur",                 "Tamil Nadu"),
    ("Pimpri Chinchwad College of Engineering",                             "PCCOE",          "Pune",                        "Maharashtra"),
    ("Prasad V Potluri Siddhartha Institute of Technology",                 "PVPSIT",         "Vijayawada",                  "Andhra Pradesh"),
    ("Presidency University, Bengaluru",                                    "",               "Bengaluru",                   "Karnataka"),
    ("Prince Shri Venkateshwara Padmavathy Engineering College",            "",               "Kancheepuram",                "Tamil Nadu"),
    ("PSNA College of Engineering and Technology, Dindigul",                "PSNA CET",       "Dindigul",                    "Tamil Nadu"),
    ("Puducherry Technological University",                                 "PTU",            "Puducherry",                  "Pondicherry"),
    ("QIS College of Engineering & Technology",                             "",               "Ongole",                      "Andhra Pradesh"),
    ("R. M. K. College of Engineering and Technology",                      "",               "Thiruvallur",                 "Tamil Nadu"),
    ("R.M.D Engineering College",                                           "",               "Thiruvallur",                 "Tamil Nadu"),
    ("Rabindranath Tagore University",                                      "RNTU",           "Raisen",                      "Madhya Pradesh"),
    ("Rajalakshmi Institute of Technology",                                 "",               "Thiruvallur",                 "Tamil Nadu"),
    ("Rajeev Gandhi Memorial College of Engineering & Technology",          "RGMCET",         "Nandyal",                     "Andhra Pradesh"),
    ("Rathinam Technical Campus",                                           "",               "Coimbatore",                  "Tamil Nadu"),
    ("Reva University",                                                     "REVA",           "Bengaluru",                   "Karnataka"),
    ("Saveetha Engineering College",                                        "",               "Sriperumbudur",               "Tamil Nadu"),
    ("Shri Vile Parle Kelavani Mandal's Dwarkadas J. Sanghvi College of Engineering", "DJSCE", "Mumbai Suburban",           "Maharashtra"),
    ("Sikkim Manipal Institute of Technology (SMIT)",                       "SMIT",           "Rangpo",                      "Sikkim"),
    ("Silicon University, Odisha, Bhubaneswar",                             "",               "Bhubaneswar",                 "Odisha"),
    ("SNS College of Technology",                                           "SNSCT",          "Coimbatore",                  "Tamil Nadu"),
    ("Sree Vidyanikethan Engineering College",                              "SVEC",           "Tirupati",                    "Andhra Pradesh"),
    ("Sri Eshwar College of Engineering",                                   "SECE",           "Coimbatore",                  "Tamil Nadu"),
    ("Sri Manakula Vinayagar Engineering College",                          "SMVEC",          "Puducherry",                  "Pondicherry"),
    ("Sri Venkateswara College of Engineering",                             "SVCE",           "Sriperumbudur",               "Tamil Nadu"),
    ("Sri Venkateswara College of Engineering and Technology",              "",               "Chittoor",                    "Andhra Pradesh"),
    ("Sri Venkateswara College of Engineering, Tirupati",                   "",               "Tirupati",                    "Andhra Pradesh"),
    ("Sri Venkateswara University",                                         "SVU",            "Tirupati",                    "Andhra Pradesh"),
    ("St. Joseph's Institute of Technology",                                "",               "Chennai",                     "Tamil Nadu"),
    ("St. Josephs College of Engineering",                                  "",               "Chennai",                     "Tamil Nadu"),
    ("Teerthanker Mahaveer University",                                     "TMU",            "Moradabad",                   "Uttar Pradesh"),
    ("The LNM Institute of Information Technology, Jaipur",                 "LNMIIT",         "Jaipur",                      "Rajasthan"),
    ("Uttaranchal University",                                              "",               "Dehradun",                    "Uttarakhand"),
    ("Vidya Jyothi Institute of Technology",                                "VJIT",           "Hyderabad",                   "Telangana"),
    ("Vignan's Institute of Information Technology",                        "VIIT",           "Visakhapatnam",               "Andhra Pradesh"),
    ("Walchand College of Engineering",                                     "WCE",            "Sangli",                      "Maharashtra"),
    ("National Institute of Food Technology, Entrepreneurship and Management - Thanjavur (NIFTEM)", "NIFTEM-T", "Thanjavur", "Tamil Nadu"),
]

# ═══════════════════════════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Rankings with Short Names"

# Styles
HEADER_BG   = "1A73E8"
HEADER_FG   = "FFFFFF"
RANKED_BG   = "EBF3FD"
RANKED_ALT  = "FFFFFF"
UNRANKED_BG = "F9FAFB"
UNRANKED_ALT= "FFFFFF"
BORDER_COL  = "D1D5DB"
SHORT_COLOR = "1557B0"   # Dark blue for short name

thin  = Side(style="thin",  color=BORDER_COL)
full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_border():
    w = Side(style="medium", color="FFFFFF")
    return Border(left=w, right=w, top=w, bottom=w)

# Column widths: Rank | College Name | Short Name | City | State
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 62
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 22

# ── Title
ws.merge_cells("A1:E1")
t = ws["A1"]
t.value = "India Rankings 2025 – Engineering Colleges with Short Names (NIRF)"
t.font = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="0F4C81")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Summary
ws.merge_cells("A2:E2")
s = ws["A2"]
total = 100 + len(band_101_150) + len(band_151_200) + len(band_201_300)
s.value = f"Total Ranked: 100   |   Total Colleges: {total}   |   Short Name = Globally recognized name only   |   Source: NIRF 2025"
s.font = Font(name="Calibri", italic=True, size=10, color="374151")
s.fill = PatternFill("solid", fgColor="DBEAFE")
s.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# ── Column headers
COLS = ["Rank", "College Name", "Short Name", "City", "State"]
for c, h in enumerate(COLS, 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = hdr_border()
ws.row_dimensions[3].height = 28

def band_sep(row, label):
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="374151")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 19

def write_row(row, rank_val, name, short, city, state, ranked=True, alt=False):
    bg = (RANKED_ALT if alt else RANKED_BG) if ranked else (UNRANKED_ALT if alt else UNRANKED_BG)
    vals = [rank_val if rank_val is not None else "", name, short, city, state]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = full_border
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center" if c == 1 else "left")
        if c == 1:
            cell.font = Font(name="Calibri", bold=True, size=10,
                             color="1A73E8" if ranked else "9CA3AF")
        elif c == 3:  # Short Name column
            cell.font = Font(name="Calibri", bold=True, size=10, color=SHORT_COLOR if v else "C0C0C0")
        else:
            cell.font = Font(name="Calibri", size=10, color="111827")
    ws.row_dimensions[row].height = 18

current_row = 4

# TOP 100
band_sep(current_row, "RANKED COLLEGES — Top 100 (Specific NIRF Rank 1–100)")
current_row += 1
for i, (rank, name, short, city, state) in enumerate(ranked):
    write_row(current_row, rank, name, short, city, state, ranked=True, alt=(i % 2 == 1))
    current_row += 1

# BAND 101-150
band_sep(current_row, "RANK BAND: 101–150")
current_row += 1
for i, (name, short, city, state) in enumerate(band_101_150):
    write_row(current_row, "101-150", name, short, city, state, ranked=False, alt=(i % 2 == 1))
    current_row += 1

# BAND 151-200
band_sep(current_row, "RANK BAND: 151–200")
current_row += 1
for i, (name, short, city, state) in enumerate(band_151_200):
    write_row(current_row, "151-200", name, short, city, state, ranked=False, alt=(i % 2 == 1))
    current_row += 1

# BAND 201-300
band_sep(current_row, "RANK BAND: 201–300")
current_row += 1
for i, (name, short, city, state) in enumerate(band_201_300):
    write_row(current_row, "201-300", name, short, city, state, ranked=False, alt=(i % 2 == 1))
    current_row += 1

ws.freeze_panes = "A4"

out = "/app/college_rankings_with_shortnames.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: 100 ranked + {len(band_101_150)} (101-150) + {len(band_151_200)} (151-200) + {len(band_201_300)} (201-300) = {total} total")

# Quick stats
with_short = sum(1 for _, s, _, _ in band_101_150 if s) + \
             sum(1 for _, s, _, _ in band_151_200 if s) + \
             sum(1 for _, s, _, _ in band_201_300 if s) + \
             sum(1 for _, _, s, _, _ in ranked if s)
print(f"Colleges WITH globally recognized short name: {with_short}")
print(f"Colleges WITHOUT short name (empty): {total - with_short}")

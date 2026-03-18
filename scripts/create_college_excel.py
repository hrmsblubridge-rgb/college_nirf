import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ─── TOP 100 RANKED COLLEGES ────────────────────────────────────────────────
ranked = [
    (1,  "Indian Institute of Technology Madras", "Chennai", "Tamil Nadu"),
    (2,  "Indian Institute of Technology Delhi", "New Delhi", "Delhi"),
    (3,  "Indian Institute of Technology Bombay", "Mumbai", "Maharashtra"),
    (4,  "Indian Institute of Technology Kanpur", "Kanpur", "Uttar Pradesh"),
    (5,  "Indian Institute of Technology Kharagpur", "Kharagpur", "West Bengal"),
    (6,  "Indian Institute of Technology Roorkee", "Roorkee", "Uttarakhand"),
    (7,  "Indian Institute of Technology Hyderabad", "Hyderabad", "Telangana"),
    (8,  "Indian Institute of Technology Guwahati", "Guwahati", "Assam"),
    (9,  "National Institute of Technology Tiruchirappalli", "Tiruchirappalli", "Tamil Nadu"),
    (10, "Indian Institute of Technology (Banaras Hindu University) Varanasi", "Varanasi", "Uttar Pradesh"),
    (11, "Birla Institute of Technology & Science - Pilani", "Pilani", "Rajasthan"),
    (12, "Indian Institute of Technology Indore", "Indore", "Madhya Pradesh"),
    (13, "National Institute of Technology Rourkela", "Rourkela", "Odisha"),
    (14, "S.R.M. Institute of Science and Technology", "Chennai", "Tamil Nadu"),
    (15, "Indian Institute of Technology (Indian School of Mines)", "Dhanbad", "Jharkhand"),
    (16, "Vellore Institute of Technology", "Vellore", "Tamil Nadu"),
    (17, "National Institute of Technology Karnataka, Surathkal", "Surathkal", "Karnataka"),
    (18, "Jadavpur University", "Kolkata", "West Bengal"),
    (19, "Indian Institute of Technology Patna", "Patna", "Bihar"),
    (20, "Anna University", "Chennai", "Tamil Nadu"),
    (21, "National Institute of Technology Calicut", "Kozhikode", "Kerala"),
    (22, "Siksha 'O' Anusandhan", "Bhubaneswar", "Odisha"),
    (23, "Amrita Vishwavidyapeetham", "Coimbatore", "Tamil Nadu"),
    (24, "Jamia Millia Islamia", "New Delhi", "Delhi"),
    (25, "Indian Institute of Technology Gandhinagar", "Gandhinagar", "Gujarat"),
    (26, "Indian Institute of Technology Mandi", "Mandi", "Himachal Pradesh"),
    (27, "Indian Institute of Technology Jodhpur", "Jodhpur", "Rajasthan"),
    (28, "National Institute of Technology Warangal", "Warangal", "Telangana"),
    (29, "Thapar Institute of Engineering and Technology", "Patiala", "Punjab"),
    (30, "Delhi Technological University", "New Delhi", "Delhi"),
    (31, "Chandigarh University", "Mohali", "Punjab"),
    (32, "Indian Institute of Technology Ropar", "Rupnagar", "Punjab"),
    (33, "Kalasalingam Academy of Research and Education", "Krishnan Koil", "Tamil Nadu"),
    (34, "Aligarh Muslim University", "Aligarh", "Uttar Pradesh"),
    (35, "Koneru Lakshmaiah Education Foundation University (K L College of Engineering)", "Vaddeswaram", "Andhra Pradesh"),
    (36, "Kalinga Institute of Industrial Technology", "Bhubaneswar", "Odisha"),
    (37, "Amity University", "Gautam Budh Nagar", "Uttar Pradesh"),
    (38, "International Institute of Information Technology Hyderabad", "Hyderabad", "Telangana"),
    (39, "Indian Institute of Technology Bhubaneswar", "Bhubaneswar", "Odisha"),
    (40, "Shanmugha Arts Science Technology & Research Academy", "Thanjavur", "Tamil Nadu"),
    (41, "Institute of Chemical Technology", "Mumbai", "Maharashtra"),
    (42, "Manipal Academy of Higher Education", "Manipal", "Karnataka"),
    (43, "UPES", "Dehradun", "Uttarakhand"),
    (44, "Visvesvaraya National Institute of Technology, Nagpur", "Nagpur", "Maharashtra"),
    (45, "Saveetha Institute of Medical and Technical Sciences", "Chennai", "Tamil Nadu"),
    (46, "Symbiosis International", "Pune", "Maharashtra"),
    (47, "Sri Sivasubramaniya Nadar College of Engineering", "Kalavakkam", "Tamil Nadu"),
    (48, "Lovely Professional University", "Phagwara", "Punjab"),
    (49, "National Institute of Technology Durgapur", "Durgapur", "West Bengal"),
    (50, "National Institute of Technology Silchar", "Silchar", "Assam"),
    (51, "Birla Institute of Technology", "Ranchi", "Jharkhand"),
    (52, "Graphic Era University", "Dehradun", "Uttarakhand"),
    (53, "National Institute of Technology Patna", "Patna", "Bihar"),
    (54, "Indian Institute of Engineering Science and Technology, Shibpur", "Howrah", "West Bengal"),
    (55, "Dr. B.R Ambedkar National Institute of Technology, Jalandhar", "Jalandhar", "Punjab"),
    (56, "Indian Institute of Technology Jammu", "Jammu", "Jammu and Kashmir"),
    (57, "Indian Institute of Technology, Tirupati", "Tirupati", "Andhra Pradesh"),
    (58, "Manipal University Jaipur", "Jaipur", "Rajasthan"),
    (59, "Manipal Institute of Technology", "Manipal", "Karnataka"),
    (60, "Madan Mohan Malaviya University of Technology", "Gorakhpur", "Uttar Pradesh"),
    (61, "Indian Institute of Space Science and Technology", "Thiruvananthapuram", "Kerala"),
    (62, "Motilal Nehru National Institute of Technology", "Prayagraj", "Uttar Pradesh"),
    (63, "Indraprastha Institute of Information Technology", "New Delhi", "Delhi"),
    (64, "Indian Institute of Technology Palakkad", "Palakkad", "Kerala"),
    (65, "National Institute of Technology Delhi", "Delhi", "Delhi"),
    (66, "Sardar Vallabhbhai National Institute of Technology", "Surat", "Gujarat"),
    (67, "PSG College of Technology", "Coimbatore", "Tamil Nadu"),
    (68, "National Institute of Technology, Andhra Pradesh", "Tadepalligudem", "Andhra Pradesh"),
    (69, "Sathyabama Institute of Science and Technology", "Chennai", "Tamil Nadu"),
    (70, "International Institute of Information Technology Bangalore", "Bengaluru", "Karnataka"),
    (71, "Netaji Subhas University of Technology (NSUT)", "Delhi", "Delhi"),
    (72, "Banasthali Vidyapith", "Banasthali", "Rajasthan"),
    (73, "Indian Institute of Technology Bhilai", "Durg", "Chhattisgarh"),
    (74, "National Institute of Technology Srinagar", "Srinagar", "Jammu and Kashmir"),
    (75, "University of Hyderabad", "Hyderabad", "Telangana"),
    (76, "M. S. Ramaiah Institute of Technology", "Bengaluru", "Karnataka"),
    (77, "Christ University", "Bengaluru", "Karnataka"),
    (78, "Indian Institute of Technology Dharwad", "Dharwad", "Karnataka"),
    (79, "Rajiv Gandhi Institute of Petroleum Technology", "Amethi", "Uttar Pradesh"),
    (80, "Sant Longowal Institute of Engineering & Technology", "Longowal", "Punjab"),
    (81, "Vignan's Foundation for Science, Technology and Research", "Guntur", "Andhra Pradesh"),
    (82, "Maulana Azad National Institute of Technology", "Bhopal", "Madhya Pradesh"),
    (83, "National Institute of Technology, Jamshedpur", "Jamshedpur", "Jharkhand"),
    (84, "National Institute of Technology Meghalaya", "Shillong", "Meghalaya"),
    (85, "Jain University, Bangalore", "Bengaluru", "Karnataka"),
    (86, "National Institute of Technology Kurukshetra", "Kurukshetra", "Haryana"),
    (87, "National Institute of Technology, Raipur", "Raipur", "Chhattisgarh"),
    (88, "Vel Tech Rangarajan Dr. Sagunthala R&D Institute of Science and Technology", "Chennai", "Tamil Nadu"),
    (89, "AU College of Engineering (A)", "Visakhapatnam", "Andhra Pradesh"),
    (90, "Chitkara University", "Rajpura", "Punjab"),
    (91, "COEP Technological University", "Pune", "Maharashtra"),
    (92, "SR University", "Warangal", "Telangana"),
    (93, "Defence Institute of Advanced Technology", "Pune", "Maharashtra"),
    (94, "Panjab University", "Chandigarh", "Chandigarh"),
    (95, "Jawaharlal Nehru Technological University", "Hyderabad", "Telangana"),
    (96, "C.V. Raman Global University, Odisha", "Bhubaneswar", "Odisha"),
    (97, "Atal Bihari Vajpayee Indian Institute of Information Technology and Management", "Gwalior", "Madhya Pradesh"),
    (98, "National Institute of Technology Hamirpur", "Hamirpur", "Himachal Pradesh"),
    (99, "Pandit Deendayal Energy University", "Gandhinagar", "Gujarat"),
    (100, "National Institute of Technology Puducherry", "Karaikal", "Puducherry"),
]

# ─── BAND 101–150 (no specific rank) ────────────────────────────────────────
band_101_150 = [
    ("Amity University", "North Twenty Four Parganas", "West Bengal"),
    ("Amity University Haryana, Gurgaon", "Gurugram", "Haryana"),
    ("Anurag University", "Hyderabad", "Telangana"),
    ("Bansilal Ramnath Agarwal Charitable Trust's Vishwakarma Institute of Technology", "Pune", "Maharashtra"),
    ("Chandigarh Engineering College-CGC, Landran, Mohali", "Mohali", "Punjab"),
    ("Chennai Institute of Technology", "Chennai", "Tamil Nadu"),
    ("Coimbatore Institute of Technology", "Coimbatore", "Tamil Nadu"),
    ("College of Engineering Trivandrum", "Thiruvananthapuram", "Kerala"),
    ("Dr. Vishwanath Karad MIT World Peace University", "Pune", "Maharashtra"),
    ("Easwari Engineering College", "Chennai", "Tamil Nadu"),
    ("Galgotias University", "Gautam Budh Nagar", "Uttar Pradesh"),
    ("Gandhi Institute of Technology And Management (GITAM)", "Visakhapatnam", "Andhra Pradesh"),
    ("Guru Gobind Singh Indraprastha University", "New Delhi", "Delhi"),
    ("Hindustan Institute of Technology and Science (HITS)", "Chennai", "Tamil Nadu"),
    ("Indian Institute of Information Technology Allahabad", "Prayagraj", "Uttar Pradesh"),
    ("Indian Institute of Technology Goa", "Ponda", "Goa"),
    ("Jawaharlal Nehru Technological University", "Kakinada", "Andhra Pradesh"),
    ("Jaypee Institute of Information Technology", "Noida", "Uttar Pradesh"),
    ("K L University", "Guntur", "Andhra Pradesh"),
    ("Karunya Institute of Technology and Sciences", "Coimbatore", "Tamil Nadu"),
    ("Kongu Engineering College", "Perundurai", "Tamil Nadu"),
    ("KPR Institute of Engineering and Technology", "Coimbatore", "Tamil Nadu"),
    ("Maharishi Markandeshwar", "Ambala", "Haryana"),
    ("Mahindra University", "Hyderabad", "Telangana"),
    ("Manav Rachna International Institute of Research & Studies", "Faridabad", "Haryana"),
    ("Maulana Azad National Urdu University", "Hyderabad", "Telangana"),
    ("National Institute for Solar Energy", "Gurugram", "Haryana"),
    ("National Institute of Food Technology, Entrepreneurship & Management", "Sonepat", "Haryana"),
    ("Nirma University", "Ahmedabad", "Gujarat"),
    ("Nitte Meenakshi Institute of Technology", "Bengaluru", "Karnataka"),
    ("Noida Institute of Engineering & Technology", "Greater Noida", "Uttar Pradesh"),
    ("Pandit Dwarka Prasad Mishra IIITDM Jabalpur", "Jabalpur", "Madhya Pradesh"),
    ("PES University", "Bengaluru", "Karnataka"),
    ("PSG Institute of Technology and Applied Research", "Coimbatore", "Tamil Nadu"),
    ("Punjab Engineering College (Deemed to be University), Chandigarh", "Chandigarh", "Chandigarh"),
    ("R.V. College of Engineering", "Bengaluru", "Karnataka"),
    ("Rajalakshmi Engineering College", "Chennai", "Tamil Nadu"),
    ("Sharda University", "Greater Noida", "Uttar Pradesh"),
    ("Shoolini University of Biotechnology and Management Sciences", "Solan", "Himachal Pradesh"),
    ("Siddaganga Institute of Technology", "Tumkur", "Karnataka"),
    ("SVKM's Narsee Monjee Institute of Management Studies", "Mumbai", "Maharashtra"),
    ("The Northcap University", "Gurugram", "Haryana"),
    ("Thiagarajar College of Engineering", "Madurai", "Tamil Nadu"),
    ("University College of Engineering", "Hyderabad", "Telangana"),
    ("University of Allahabad", "Prayagraj", "Uttar Pradesh"),
    ("Veermata Jijabai Technological Institute (VJTI, Mumbai)", "Mumbai", "Maharashtra"),
    ("National Institute of Technology Agartala", "Agartala", "Tripura"),
    ("National Institute of Technology Arunachal Pradesh", "Itanagar", "Arunachal Pradesh"),
    ("National Institute of Technology Goa", "Cuncolim", "Goa"),
    ("National Institute of Technology Mizoram", "Aizawl", "Mizoram"),
]

# ─── BAND 151–200 (no specific rank) ────────────────────────────────────────
band_151_200 = [
    ("Amity University Rajasthan, Jaipur", "Jaipur", "Rajasthan"),
    ("B. S. Abdur Rahman Crescent Institute of Science and Technology", "Chennai", "Tamil Nadu"),
    ("B.M.S. College of Engineering", "Bengaluru", "Karnataka"),
    ("Bharati Vidyapeeth Deemed University College of Engineering", "Pune", "Maharashtra"),
    ("C M R Institute of Technology", "Bengaluru", "Karnataka"),
    ("CGC College of Engineering, Landran", "Sahibzada Ajit Singh Nagar", "Punjab"),
    ("Chaitanya Bharathi Institute of Technology", "Hyderabad", "Telangana"),
    ("CVR College of Engineering", "Ibrahimpatam", "Telangana"),
    ("Dayalbagh Educational Institute", "Agra", "Uttar Pradesh"),
    ("Dr. D. Y. Patil Institute of Technology", "Pune", "Maharashtra"),
    ("Dr. M. G. R. Educational and Research Institute", "Chennai", "Tamil Nadu"),
    ("G. L. A. University", "Mathura", "Uttar Pradesh"),
    ("G.L. Bajaj Institute of Technology and Management", "Greater Noida", "Uttar Pradesh"),
    ("Goka Raju Ranga Raju Institute of Engineering & Technology", "Hyderabad", "Telangana"),
    ("Guru Ghasidas Vishwavidyalaya", "Bilaspur", "Chhattisgarh"),
    ("Indian Institute of Information Technology, Design & Manufacturing, Kancheepuram", "Chennai", "Tamil Nadu"),
    ("Institute of Aeronautical Engineering", "Hyderabad", "Telangana"),
    ("Institute of Engineering & Management", "Kolkata", "West Bengal"),
    ("Integral University", "Lucknow", "Uttar Pradesh"),
    ("Islamic University of Science & Technology, Pulwama", "Pulwama", "Jammu and Kashmir"),
    ("J. C. Bose University of Science and Technology, YMCA", "Faridabad", "Haryana"),
    ("Jawaharlal Nehru Technological University, Ananthapuramu", "Ananthapuramu", "Andhra Pradesh"),
    ("JSPM's Rajarshi Shahu College of Engineering", "Pune", "Maharashtra"),
    ("KIET Group of Institutions", "Ghaziabad", "Uttar Pradesh"),
    ("KLE Technological University", "Dharwad", "Karnataka"),
    ("Kumaraguru College of Technology", "Coimbatore", "Tamil Nadu"),
    ("Maulana Abul Kalam Azad University of Technology", "Nadia", "West Bengal"),
    ("Mepco Schlenk Engineering College", "Sivakasi", "Tamil Nadu"),
    ("National Institute of Technology Manipur", "Imphal", "Manipur"),
    ("National Institute of Technology Sikkim", "South Sikkim", "Sikkim"),
    ("National Institute of Technology Uttarakhand", "Srinagar (Garhwal)", "Uttarakhand"),
    ("National Institute of Technology Nagaland", "Dimapur", "Nagaland"),
    ("New Horizon College of Engineering", "Bengaluru", "Karnataka"),
    ("NMAM Institute of Technology", "Nitte, Udupi", "Karnataka"),
    ("North Eastern Regional Institute of Science & Technology", "Itanagar", "Arunachal Pradesh"),
    ("R.M.K. Engineering College", "Thiruvallur", "Tamil Nadu"),
    ("Ramdeobaba University, Nagpur", "Nagpur", "Maharashtra"),
    ("Shri Mata Vaishno Devi University", "Katra", "Jammu and Kashmir"),
    ("Sona College of Technology", "Salem", "Tamil Nadu"),
    ("Sri Ramakrishna Engineering College", "Coimbatore", "Tamil Nadu"),
    ("Sri Sai Ram Institute of Technology", "Chennai", "Tamil Nadu"),
    ("Sri Sairam Engineering College", "Kancheepuram", "Tamil Nadu"),
    ("Tezpur University", "Tezpur", "Assam"),
    ("The National Institute of Engineering", "Mysuru", "Karnataka"),
    ("Vallurupalli Nageswara Rao Vignana Jyothi Institute of Engineering and Technology", "Hyderabad", "Telangana"),
    ("Vardhaman College of Engineering", "Rangareddy", "Telangana"),
    ("Veer Surendra Sai University of Technology", "Sambalpur", "Odisha"),
    ("Velagapudi Ramakrishna Siddhartha Engineering College", "Vijayawada", "Andhra Pradesh"),
    ("Vels Institute of Science Technology and Advanced Studies (VISTAS)", "Chennai", "Tamil Nadu"),
    ("Vignan Institute of Technology and Science", "Yadadri-Bhuvanagiri", "Telangana"),
    ("Visvesvaraya Technological University", "Belgaum", "Karnataka"),
    ("Yeshwantrao Chavan College of Engineering", "Nagpur", "Maharashtra"),
]

# ─── BAND 201–300 (no specific rank) ────────────────────────────────────────
band_201_300 = [
    ("Aditya Institute of Technology and Management", "Tekkali", "Andhra Pradesh"),
    ("Aditya University", "Surampalem", "Andhra Pradesh"),
    ("Ajeenkya D Y Patil University", "Pune", "Maharashtra"),
    ("Amity University Patna", "Patna", "Bihar"),
    ("Amity University, Gwalior", "Gwalior", "Madhya Pradesh"),
    ("Amity University, Jharkhand", "Ranchi", "Jharkhand"),
    ("Annamalai University", "Annamalainagar", "Tamil Nadu"),
    ("Army Institute of Technology", "Pune", "Maharashtra"),
    ("B I T SINDRI", "Dhanbad", "Jharkhand"),
    ("BML Munjal University", "Gurgaon", "Haryana"),
    ("BMS Institute of Technology & Management", "Bengaluru", "Karnataka"),
    ("Centurion University of Technology and Management", "Paralakhemundi", "Odisha"),
    ("Chandigarh Engineering College Jhanjeri", "Sahibzada Ajit Singh Nagar", "Punjab"),
    ("CMR College of Engineering & Technology", "Hyderabad", "Telangana"),
    ("CMR Technical Campus", "Hyderabad", "Telangana"),
    ("College of Engineering & Technology, Bhubaneswar", "Bhubaneswar", "Odisha"),
    ("Dayananda Sagar College of Engineering", "Bengaluru", "Karnataka"),
    ("Dhirubhai Ambani Institute of Information and Communication Technology", "Gandhinagar", "Gujarat"),
    ("Dr. Shyama Prasad Mukherjee International Institute of Information Technology, Naya Raipur", "Raipur", "Chhattisgarh"),
    ("E.G.S. Pillay Engineering College", "Nagapattinam", "Tamil Nadu"),
    ("G. H. Raisoni College of Engineering, Nagpur", "Nagpur", "Maharashtra"),
    ("G. Narayanamma Institute of Technology & Science for Women", "Hyderabad", "Telangana"),
    ("Galgotias College of Engineering & Technology", "Greater Noida", "Uttar Pradesh"),
    ("Gandhi Institute for Technological Advancement (GITA), Bhubaneswar", "Bhubaneswar", "Odisha"),
    ("GIET University, Gunupur", "Gunupur", "Odisha"),
    ("GMR Institute of Technology", "Rajahmundry", "Andhra Pradesh"),
    ("Godavari Institute of Engineering & Technology", "Rajahmundry", "Andhra Pradesh"),
    ("Guru Jambheshwar University of Science and Technology, Hisar", "Hisar", "Haryana"),
    ("Harcourt Butler Technical University", "Kanpur Nagar", "Uttar Pradesh"),
    ("Hindusthan College of Engineering and Technology", "Coimbatore", "Tamil Nadu"),
    ("IES College of Technology, Bhopal", "Bhopal", "Madhya Pradesh"),
    ("Indian Institute of Information Technology Guwahati", "Guwahati", "Assam"),
    ("Indian Institute of Petroleum & Energy", "Visakhapatnam", "Andhra Pradesh"),
    ("Indira Gandhi Delhi Technical University for Women", "Delhi", "Delhi"),
    ("Jamia Hamdard", "New Delhi", "Delhi"),
    ("Jaypee University of Information Technology", "Solan", "Himachal Pradesh"),
    ("JIS College of Engineering", "Kalyani", "West Bengal"),
    ("JSS Academy Of Technical Education, Noida", "Gautam Budh Nagar", "Uttar Pradesh"),
    ("JSS Science and Technology University", "Mysuru", "Karnataka"),
    ("K. Ramakrishnan College of Engineering", "Samayapuram", "Tamil Nadu"),
    ("K. Ramakrishnan College of Technology", "Tiruchirappalli", "Tamil Nadu"),
    ("Kakatiya Institute of Technology & Science", "Warangal", "Telangana"),
    ("Kalaignar Karunanidhi Institute of Technology", "Coimbatore", "Tamil Nadu"),
    ("Karpagam College of Engineering", "Coimbatore", "Tamil Nadu"),
    ("Lakshmi Narain College of Technology", "Bhopal", "Madhya Pradesh"),
    ("M. Kumarasamy College of Engineering", "Karur", "Tamil Nadu"),
    ("Madanapalle Institute of Technology & Science", "Madanapalle", "Andhra Pradesh"),
    ("Maharaja Sayajirao University of Baroda", "Vadodara", "Gujarat"),
    ("Maharshi Dayanand University, Rohtak", "Rohtak", "Haryana"),
    ("Maharshi Karve Stree Shikshan Samstha's Cummins College of Engineering for Women", "Pune", "Maharashtra"),
    ("Mahatma Jyotiba Phule Rohilkhand University, Bareilly", "Bareilly", "Uttar Pradesh"),
    ("Malla Reddy Engineering College", "Hyderabad", "Telangana"),
    ("Malla Reddy Engineering College for Women (Autonomous)", "Hyderabad", "Telangana"),
    ("Manav Rachna University", "Faridabad", "Haryana"),
    ("Marwadi University", "Rajkot", "Gujarat"),
    ("MIT Art, Design and Technology University, Pune", "Pune", "Maharashtra"),
    ("MLR Institute of Technology", "Hyderabad", "Telangana"),
    ("Narula Institute of Technology", "Kolkata", "West Bengal"),
    ("National Engineering College", "Kovilpatti", "Tamil Nadu"),
    ("National Institute of Advanced Manufacturing Technology, Ranchi", "Ranchi", "Jharkhand"),
    ("National Institute Of Technology, Andhra Pradesh", "Tadepalligudem", "Andhra Pradesh"),
    ("P E S College of Engineering, MANDYA", "Mandya", "Karnataka"),
    ("Padmashree Dr. D.Y. Patil Vidyapeeth, Mumbai", "Mumbai", "Maharashtra"),
    ("Panimalar Engineering College", "Thiruvallur", "Tamil Nadu"),
    ("Pimpri Chinchwad College of Engineering", "Pune", "Maharashtra"),
    ("Prasad V Potluri Siddhartha Institute of Technology", "Vijayawada", "Andhra Pradesh"),
    ("Presidency University, Bengaluru", "Bengaluru", "Karnataka"),
    ("Prince Shri Venkateshwara Padmavathy Engineering College", "Kancheepuram", "Tamil Nadu"),
    ("PSNA College of Engineering and Technology, Dindigul", "Dindigul", "Tamil Nadu"),
    ("Puducherry Technological University", "Puducherry", "Pondicherry"),
    ("QIS College of Engineering & Technology", "Ongole", "Andhra Pradesh"),
    ("R. M. K. College of Engineering and Technology", "Thiruvallur", "Tamil Nadu"),
    ("R.M.D Engineering College", "Thiruvallur", "Tamil Nadu"),
    ("Rabindranath Tagore University", "Raisen", "Madhya Pradesh"),
    ("Rajalakshmi Institute of Technology", "Thiruvallur", "Tamil Nadu"),
    ("Rajeev Gandhi Memorial College of Engineering & Technology", "Nandyal", "Andhra Pradesh"),
    ("Rathinam Technical Campus", "Coimbatore", "Tamil Nadu"),
    ("Reva University", "Bengaluru", "Karnataka"),
    ("Saveetha Engineering College", "Sriperumbudur", "Tamil Nadu"),
    ("Shri Vile Parle Kelavani Mandal's Dwarkadas J. Sanghvi College of Engineering", "Mumbai Suburban", "Maharashtra"),
    ("Sikkim Manipal Institute of Technology (SMIT)", "Rangpo", "Sikkim"),
    ("Silicon University, Odisha, Bhubaneswar", "Bhubaneswar", "Odisha"),
    ("SNS College of Technology", "Coimbatore", "Tamil Nadu"),
    ("Sree Vidyanikethan Engineering College", "Tirupati", "Andhra Pradesh"),
    ("Sri Eshwar College of Engineering", "Coimbatore", "Tamil Nadu"),
    ("Sri Manakula Vinayagar Engineering College", "Puducherry", "Pondicherry"),
    ("Sri Venkateswara College of Engineering", "Sriperumbudur", "Tamil Nadu"),
    ("Sri Venkateswara College of Engineering and Technology", "Chittoor", "Andhra Pradesh"),
    ("Sri Venkateswara College of Engineering, Tirupati", "Tirupati", "Andhra Pradesh"),
    ("Sri Venkateswara University", "Tirupati", "Andhra Pradesh"),
    ("St. Joseph's Institute of Technology", "Chennai", "Tamil Nadu"),
    ("St. Josephs College of Engineering", "Chennai", "Tamil Nadu"),
    ("Teerthanker Mahaveer University", "Moradabad", "Uttar Pradesh"),
    ("The LNM Institute of Information Technology, Jaipur", "Jaipur", "Rajasthan"),
    ("Uttaranchal University", "Dehradun", "Uttarakhand"),
    ("Vidya Jyothi Institute of Technology", "Hyderabad", "Telangana"),
    ("Vignan's Institute of Information Technology", "Visakhapatnam", "Andhra Pradesh"),
    ("Walchand College of Engineering", "Sangli", "Maharashtra"),
    ("National Institute of Food Technology, Entrepreneurship and Management - Thanjavur (NIFTEM)", "Thanjavur", "Tamil Nadu"),
]

# ─── BUILD WORKBOOK ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "College Rankings"

# Colors
HEADER_BG   = "1A73E8"   # Blue
HEADER_FG   = "FFFFFF"   # White
RANKED_BG   = "EBF3FD"   # Light blue tint
RANKED_ALT  = "FFFFFF"   # White
UNRANKED_BG = "F9FAFB"   # Very light gray
UNRANKED_ALT= "FFFFFF"   # White
BORDER_COL  = "D1D5DB"   # Gray border
BAND_HEADER = "374151"   # Dark gray
BAND_BG     = "F3F4F6"   # Light gray for band separator

thin  = Side(style="thin",   color=BORDER_COL)
thick = Side(style="medium", color="1A73E8")
full_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

def hdr_border():
    return Border(left=Side(style="medium", color="FFFFFF"),
                  right=Side(style="medium", color="FFFFFF"),
                  top=Side(style="medium", color="FFFFFF"),
                  bottom=Side(style="medium", color="FFFFFF"))

# ── Column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 25

# ── Title row
ws.merge_cells("A1:D1")
title_cell = ws["A1"]
title_cell.value = "India Rankings 2025 – Engineering Colleges (NIRF)"
title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="0F4C81")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 38

# ── Summary row
ws.merge_cells("A2:D2")
summary = ws["A2"]
summary.value = f"Total Ranked Colleges: 100   |   Total Colleges: {100 + len(band_101_150) + len(band_151_200) + len(band_201_300)}   |   Source: NIRF 2025"
summary.font = Font(name="Calibri", italic=True, size=11, color="374151")
summary.fill = PatternFill("solid", fgColor="DBEAFE")
summary.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# ── Column Headers (row 3)
headers = ["Rank", "College Name", "City", "State"]
for col, hdr in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=hdr)
    cell.font = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = hdr_border()
ws.row_dimensions[3].height = 28

# ── Helper to write a band separator row
def band_separator(row, label):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="374151")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

# ── Helper to write a data row
def write_row(row, rank_val, name, city, state, is_ranked=True, alt=False):
    bg = (RANKED_ALT if alt else RANKED_BG) if is_ranked else (UNRANKED_ALT if alt else UNRANKED_BG)
    rank_display = rank_val if rank_val is not None else ""
    values = [rank_display, name, city, state]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = full_border
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if col == 1 else "left")
        if col == 1:
            c.font = Font(name="Calibri", bold=True, size=10,
                          color="1A73E8" if is_ranked else "9CA3AF")
        else:
            c.font = Font(name="Calibri", size=10, color="111827")
    ws.row_dimensions[row].height = 18

# ── Write TOP 100
current_row = 4
band_separator(current_row, "RANKED COLLEGES — Top 100 (Specific NIRF Rank)")
current_row += 1

for i, (rank, name, city, state) in enumerate(ranked):
    write_row(current_row, rank, name, city, state, is_ranked=True, alt=(i % 2 == 1))
    current_row += 1

# ── Write BAND 101–150
band_separator(current_row, "RANK BAND: 101–150")
current_row += 1
for i, (name, city, state) in enumerate(band_101_150):
    write_row(current_row, "101-150", name, city, state, is_ranked=False, alt=(i % 2 == 1))
    current_row += 1

# ── Write BAND 151–200
band_separator(current_row, "RANK BAND: 151–200")
current_row += 1
for i, (name, city, state) in enumerate(band_151_200):
    write_row(current_row, "151-200", name, city, state, is_ranked=False, alt=(i % 2 == 1))
    current_row += 1

# ── Write BAND 201–300
band_separator(current_row, "RANK BAND: 201–300")
current_row += 1
for i, (name, city, state) in enumerate(band_201_300):
    write_row(current_row, "201-300", name, city, state, is_ranked=False, alt=(i % 2 == 1))
    current_row += 1

# ── Freeze panes
ws.freeze_panes = "A4"

# ── Save
out_path = "/app/college_rankings.xlsx"
wb.save(out_path)

total = 100 + len(band_101_150) + len(band_151_200) + len(band_201_300)
print(f"Excel created: {out_path}")
print(f"Total rows: {total}")
print(f"  - Ranked (1-100): 100")
print(f"  - Band 101-150: {len(band_101_150)}")
print(f"  - Band 151-200: {len(band_151_200)}")
print(f"  - Band 201-300: {len(band_201_300)}")
